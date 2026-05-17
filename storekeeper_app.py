"""
Asset Manager — Storekeeper App
For: Storekeeper User
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests, json, os, sys, threading
from datetime import datetime, date

# ─── Config ───────────────────────────────────────────────────────────────────

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

DEFAULT_CFG = {
    "server_url": "http://asset-server:8080"
}

def load_cfg():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH) as f:
            d = json.load(f)
        DEFAULT_CFG.update(d)
    return DEFAULT_CFG.copy()

def save_cfg(cfg):
    with open(CONFIG_PATH, "w") as f:
        json.dump(cfg, f, indent=2)

TEAM_MEMBERS = [
    "Admin User",
    "Engineer3",
    "Engineer1",
    "Engineer2",
]

ASSET_TYPES = ["Laptop", "Desktop", "Mobile", "Tablet", "Screen", "UPS", "Server",
               "Cisco Phone", "Printer", "Scanner", "Switch"]

# ─── Colours ──────────────────────────────────────────────────────────────────

BG        = "#1a2332"
CARD      = "#1e2d3d"
CARD2     = "#243447"
ACCENT    = "#0078d4"
ACCENT2   = "#106ebe"
TEXT      = "#e8edf2"
SUBTEXT   = "#7a8fa3"
SUCCESS   = "#4caf50"
WARNING   = "#f0a500"
DANGER    = "#f44336"
INPUT_BG  = "#2a3f55"

# ─── Shared API helper ────────────────────────────────────────────────────────

def api(method, url, **kw):
    kw.setdefault("timeout", 7)
    return getattr(requests, method)(url, **kw)

# ─── Date Range Export Dialog ─────────────────────────────────────────────────

class ExportDialog:
    def __init__(self, root, server_url):
        self.server_url = server_url

        win = tk.Toplevel(root)
        win.title("Export to Excel")
        win.geometry("400x420")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="📊  Export to Excel",
                 bg=BG, fg=TEXT,
                 font=("Segoe UI", 13, "bold")).pack(pady=(16, 4), padx=20, anchor="w")

        frm = tk.Frame(win, bg=CARD2, padx=20, pady=16)
        frm.pack(fill="x", padx=20, pady=8)

        tk.Label(frm, text="DATE RANGE",
                 bg=CARD2, fg=ACCENT,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 10))

        date_row = tk.Frame(frm, bg=CARD2)
        date_row.pack(fill="x")

        def date_col(label, default, side, pad):
            col = tk.Frame(date_row, bg=CARD2)
            col.pack(side=side, fill="x", expand=True, padx=pad)
            tk.Label(col, text=label, bg=CARD2, fg=SUBTEXT,
                     font=("Segoe UI", 8)).pack(anchor="w")
            e = tk.Entry(col, bg=INPUT_BG, fg=TEXT, relief="flat",
                         font=("Segoe UI", 10), insertbackground=TEXT)
            e.insert(0, default)
            e.pack(fill="x", ipady=7)
            return e

        today    = date.today()
        first    = today.replace(day=1).strftime("%d/%m/%Y")
        todaystr = today.strftime("%d/%m/%Y")

        self.from_e = date_col("From (dd/mm/yyyy)", first,    "left",  (0, 6))
        self.to_e   = date_col("To   (dd/mm/yyyy)", todaystr, "right", (6, 0))

        tk.Label(frm, text="Leave both blank to export all records.",
                 bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(8, 0))

        # Filter by member
        tk.Label(frm, text="Member  (optional)",
                 bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(10, 2))
        self.member_cb = ttk.Combobox(frm, values=["All"] + TEAM_MEMBERS,
                                      state="readonly", font=("Segoe UI", 9))
        self.member_cb.set("All")
        self.member_cb.pack(fill="x", ipady=4)

        self.err_lbl = tk.Label(win, text="", bg=BG, fg=DANGER,
                                font=("Segoe UI", 9))
        self.err_lbl.pack(pady=(4, 0))

        tk.Button(win, text="Choose File & Export",
                  bg=SUCCESS, fg="white",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", pady=10,
                  command=self._export,
                  activebackground="#388e3c").pack(fill="x", padx=20, pady=8)

        self.win = win

    def _parse_date(self, text):
        text = text.strip()
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt)
            except: pass
        return None

    def _export(self):
        from_dt = self._parse_date(self.from_e.get())
        to_dt   = self._parse_date(self.to_e.get())
        from_str = self.from_e.get().strip()
        to_str   = self.to_e.get().strip()

        if from_str and from_dt is None:
            self.err_lbl.config(text="Invalid 'From' date. Use dd/mm/yyyy")
            return
        if to_str and to_dt is None:
            self.err_lbl.config(text="Invalid 'To' date. Use dd/mm/yyyy")
            return

        # Extend to_dt to end of day
        if to_dt:
            to_dt = to_dt.replace(hour=23, minute=59, second=59)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing Library",
                "Run:  pip install openpyxl  then restart.")
            return

        try:
            r = api("get", f"{self.server_url}/export")
            if not r.ok:
                self.err_lbl.config(text="Server error.")
                return
            data = r.json()
        except Exception as e:
            self.err_lbl.config(text=f"Cannot reach server: {e}")
            return

        # Apply filters
        member_filter = self.member_cb.get()
        filtered = []
        for t in data:
            if member_filter != "All" and t['team_member'] != member_filter:
                continue
            ts = t.get('timestamp', '')
            try:
                dt = datetime.fromisoformat(ts)
                if from_dt and dt < from_dt: continue
                if to_dt   and dt > to_dt:   continue
            except: pass
            filtered.append(t)

        if not filtered:
            self.err_lbl.config(text="No records match the selected filters.")
            return

        # File dialog
        label_from = from_dt.strftime("%Y%m%d") if from_dt else "all"
        label_to   = to_dt.strftime("%Y%m%d")   if to_dt   else "now"
        default_fn = f"Assets_{label_from}_to_{label_to}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_fn,
            title="Save Excel Report"
        )
        if not path:
            return

        self.win.destroy()
        _write_excel(filtered, path, from_dt, to_dt)

# ─── Excel Writer ─────────────────────────────────────────────────────────────

def _write_excel(data, path, from_dt=None, to_dt=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Transactions"

    blue_fill   = PatternFill("solid", fgColor="0078D4")
    green_fill  = PatternFill("solid", fgColor="C8E6C9")
    red_fill    = PatternFill("solid", fgColor="FFCDD2")
    orange_fill = PatternFill("solid", fgColor="FFF3E0")
    grey_fill   = PatternFill("solid", fgColor="F5F5F5")

    hdr_font   = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
    norm_font  = Font(name="Segoe UI", size=9)
    bold_font  = Font(bold=True, name="Segoe UI", size=9)
    thin       = Side(style="thin", color="DDDDDD")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    period = ""
    if from_dt or to_dt:
        f = from_dt.strftime("%d/%m/%Y") if from_dt else "—"
        t = to_dt.strftime("%d/%m/%Y")   if to_dt   else "—"
        period = f"  |  Period: {f}  →  {t}"
    tc.value     = f"Asset Manager — Transaction Report  ({datetime.now().strftime('%B %Y')}){period}"
    tc.font      = Font(bold=True, name="Segoe UI", size=13, color="0078D4")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:L2")
    ws["A2"].value     = f"Generated: {datetime.now().strftime('%d/%m/%Y  %H:%M')}    Records: {len(data)}"
    ws["A2"].font      = Font(name="Segoe UI", size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Headers — Asset No. and Serial No. are first
    headers    = ["#", "Asset No.", "Serial No.", "Model", "Type", "Team Member",
                  "Direction", "Status", "Rejection Reason",
                  "Logged At", "Confirmed At", "Notes"]
    col_widths = [6, 14, 14, 26, 12, 22, 12, 14, 28, 18, 18, 26]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = blue_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    def fmt_ts(v):
        try: return datetime.fromisoformat(v).strftime("%d/%m/%Y  %H:%M")
        except: return v or ""

    for ri, t in enumerate(data, 4):
        if   t.get('confirmed'): status = "Confirmed"; fill = green_fill
        elif t.get('rejected'):  status = "Rejected";  fill = red_fill
        else:                    status = "Pending";   fill = orange_fill

        values = [
            t['id'],
            t.get('asset_number',     ''),
            t.get('serial_number',    ''),
            t.get('asset_model',      ''),
            t.get('asset_type',       ''),
            t.get('team_member',      ''),
            t.get('direction',        ''),
            status,
            t.get('rejection_reason', ''),
            fmt_ts(t.get('timestamp',    '')),
            fmt_ts(t.get('confirmed_at', '')),
            t.get('notes', ''),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = bold_font if ci == 1 else norm_font
            cell.alignment = center if ci in (1, 5, 6, 7, 8) else \
                             Alignment(vertical="center")
            cell.border    = border
            if ci == 8:
                cell.fill = fill
            elif ri % 2 == 0:
                cell.fill = grey_fill
        ws.row_dimensions[ri].height = 20

    # ── Summary sheet ─────────────────────────────────────────────────────────
    try:
        from openpyxl.chart import BarChart, PieChart, Reference
        HAS_CHART = True
    except ImportError:
        HAS_CHART = False

    ws2 = wb.create_sheet("Summary")
    for col, w in zip("ABCDEFG", [26, 14, 14, 14, 4, 26, 14]):
        ws2.column_dimensions[col].width = w

    total     = len(data)
    confirmed = sum(1 for t in data if t.get('confirmed'))
    rejected  = sum(1 for t in data if t.get('rejected'))
    pending   = total - confirmed - rejected
    out_count = sum(1 for t in data if t['direction'] == 'Out Store')
    in_count  = sum(1 for t in data if t['direction'] == 'In Store')

    thin_s  = Side(style="thin", color="CCCCCC")
    thin_b  = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    ctr     = Alignment(horizontal="center", vertical="center")

    # Title
    ws2.merge_cells("A1:G1")
    tc2 = ws2["A1"]
    tc2.value     = f"Asset Manager — Summary Report  ({datetime.now().strftime('%B %Y')})"
    tc2.font      = Font(bold=True, name="Segoe UI", size=14, color="0078D4")
    tc2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:G2")
    ws2["A2"].value     = f"Generated: {datetime.now().strftime('%d/%m/%Y  %H:%M')}    Total Records: {total}"
    ws2["A2"].font      = Font(name="Segoe UI", size=9, color="888888")
    ws2["A2"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[2].height = 18

    def s2_hdr(row, col, text, bg="0078D4"):
        c = ws2.cell(row=row, column=col, value=text)
        c.font      = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = ctr
        c.border    = thin_b
        ws2.row_dimensions[row].height = 22

    def s2_val(row, col, value, color="000000", bold=False, centered=False):
        c = ws2.cell(row=row, column=col, value=value)
        c.font      = Font(name="Segoe UI", size=10, bold=bold, color=color)
        c.alignment = Alignment(horizontal="center" if centered else "left", vertical="center")
        c.border    = thin_b
        ws2.row_dimensions[row].height = 20

    # Status table (A4:C8)
    s2_hdr(4, 1, "Status");  s2_hdr(4, 2, "Count");  s2_hdr(4, 3, "% of Total")
    for i, (label, val, col) in enumerate([
        ("✅ Confirmed", confirmed, "2E7D32"),
        ("⏳ Pending",   pending,   "E65100"),
        ("❌ Rejected",  rejected,  "C62828"),
        ("📊 Total",     total,     "0078D4"),
    ], 5):
        s2_val(i, 1, label, col, bold=True)
        s2_val(i, 2, val,   col, bold=True, centered=True)
        s2_val(i, 3, f"{val/total*100:.1f}%" if total else "—", col, centered=True)

    # Direction table (A11:C13)
    s2_hdr(11, 1, "Direction");  s2_hdr(11, 2, "Count");  s2_hdr(11, 3, "% of Total")
    for i, (label, val, col) in enumerate([
        ("🔴 Out Store", out_count, "C62828"),
        ("🟢 In Store",  in_count,  "2E7D32"),
    ], 12):
        s2_val(i, 1, label, col, bold=True)
        s2_val(i, 2, val,   col, bold=True, centered=True)
        s2_val(i, 3, f"{val/total*100:.1f}%" if total else "—", col, centered=True)

    # Per-member table (F4:G4+)
    s2_hdr(4, 6, "Team Member");  s2_hdr(4, 7, "Transactions")
    member_counts = []
    for i, m in enumerate(TEAM_MEMBERS, 5):
        cnt = sum(1 for t in data if t['team_member'] == m)
        member_counts.append(cnt)
        s2_val(i, 6, m)
        s2_val(i, 7, cnt, "0078D4", bold=True, centered=True)

    # Charts
    if HAS_CHART and total > 0:
        # Status bar chart
        chart1 = BarChart()
        chart1.type    = "col"
        chart1.title   = "Transaction Status"
        chart1.y_axis.title = "Count"
        chart1.style   = 10
        chart1.width   = 14
        chart1.height  = 10
        data1  = Reference(ws2, min_col=2, min_row=4, max_row=7)
        cats1  = Reference(ws2, min_col=1, min_row=5, max_row=7)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        ws2.add_chart(chart1, "A15")

        # Per-member bar chart
        chart2 = BarChart()
        chart2.type    = "col"
        chart2.title   = "Transactions per Engineer"
        chart2.y_axis.title = "Count"
        chart2.style   = 10
        chart2.width   = 14
        chart2.height  = 10
        n = len(TEAM_MEMBERS)
        data2  = Reference(ws2, min_col=7, min_row=4, max_row=4 + n)
        cats2  = Reference(ws2, min_col=6, min_row=5, max_row=4 + n)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws2.add_chart(chart2, "F15")

        # Direction pie chart
        chart3 = PieChart()
        chart3.title  = "In vs Out"
        chart3.style  = 10
        chart3.width  = 10
        chart3.height = 10
        data3  = Reference(ws2, min_col=2, min_row=11, max_row=12)
        cats3  = Reference(ws2, min_col=1, min_row=12, max_row=13)
        chart3.add_data(data3, titles_from_data=False)
        chart3.set_categories(cats3)
        ws2.add_chart(chart3, "A31")

    wb.save(path)
    messagebox.showinfo("Export Complete ✓",
        f"Saved to:\n{path}\n\n  • {total} records exported")


# ─── Main App ─────────────────────────────────────────────────────────────────

class StorekeeperApp:
    def __init__(self, root):
        self.root       = root
        self.cfg        = load_cfg()
        self.server_url = self.cfg.get("server_url", "http://asset-server:8080")
        self._all_rows  = []
        self._sort_col  = None
        self._sort_rev       = False
        self._lookup_after_id = None   # pending debounce call id
        self._lookup_locked   = False  # True when type was auto-set
        self._bulk_mode       = False  # bulk scan mode toggle
        self._bulk_queue      = []     # list of dicts queued for bulk submit
        self._server_check_timer = None  # track after() timer to prevent leaks

        self.root.title("Asset Manager — Storekeeper")
        self.root.geometry("1120x740")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.root.minsize(900, 580)

        self._setup_styles()
        self._build_ui()
        self._check_server()
        self._refresh_transactions()

    def _setup_styles(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("Treeview",
                    background=INPUT_BG, foreground=TEXT,
                    fieldbackground=INPUT_BG, rowheight=30,
                    font=("Segoe UI", 9))
        s.configure("Treeview.Heading",
                    background=ACCENT, foreground="white",
                    font=("Segoe UI", 9, "bold"))
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", "white")])

    def _build_ui(self):
        top = tk.Frame(self.root, bg=ACCENT, height=52)
        top.pack(fill="x")
        top.pack_propagate(False)
        tk.Label(top, text="📦  Asset Manager",
                 bg=ACCENT, fg="white",
                 font=("Segoe UI", 15, "bold")).pack(side="left", padx=20, pady=12)
        self.status_lbl = tk.Label(top, text="● Connecting…",
                                   bg=ACCENT, fg="#cce4ff", font=("Segoe UI", 9))
        self.status_lbl.pack(side="right", padx=20)
        tk.Label(top, text="Storekeeper: Storekeeper User",
                 bg=ACCENT, fg="#cce4ff",
                 font=("Segoe UI", 9)).pack(side="right")

        body = tk.Frame(self.root, bg=BG)
        body.pack(fill="both", expand=True, padx=16, pady=12)
        self._build_form(body)
        self._build_history(body)

    # ── Form ──────────────────────────────────────────────────────────────────

    def _build_form(self, parent):
        panel = tk.Frame(parent, bg=CARD, width=316)
        panel.pack(side="left", fill="y", padx=(0, 12))
        panel.pack_propagate(False)

        tk.Label(panel, text="New Transaction",
                 bg=CARD, fg=ACCENT,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=18, pady=(16, 4))
        tk.Frame(panel, bg=ACCENT, height=2).pack(fill="x", padx=18, pady=(0, 12))

        frm = tk.Frame(panel, bg=CARD, padx=18)
        frm.pack(fill="x")

        def lbl(text, required=False):
            f = tk.Frame(frm, bg=CARD)
            f.pack(fill="x", pady=(9, 2))
            tk.Label(f, text=text, bg=CARD, fg=SUBTEXT,
                     font=("Segoe UI", 9)).pack(side="left")
            if required:
                tk.Label(f, text=" *", bg=CARD, fg=DANGER,
                         font=("Segoe UI", 9, "bold")).pack(side="left")

        def ent(font_size=10):
            e = tk.Entry(frm, bg=INPUT_BG, fg=TEXT, relief="flat",
                         font=("Segoe UI", font_size), insertbackground=TEXT,
                         highlightthickness=1,
                         highlightbackground="#334d66",
                         highlightcolor=ACCENT)
            e.pack(fill="x", ipady=8)
            return e

        # ── Asset No. / Serial No. — TOP PRIORITY ────────────────────────────
        lbl("Asset ID  (fill at least one)", required=True)

        id_row = tk.Frame(frm, bg=CARD)
        id_row.pack(fill="x")

        for side, attr, label, pad in [
            ("left",  "v_asset_no", "Asset No.",  (0, 5)),
            ("right", "v_serial",   "Serial No.", (5, 0)),
        ]:
            col = tk.Frame(id_row, bg=CARD)
            col.pack(side=side, fill="x", expand=True, padx=pad)
            tk.Label(col, text=label, bg=CARD, fg="#0078d4",
                     font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 3))
            e = tk.Entry(col, bg=INPUT_BG, fg=TEXT, relief="flat",
                         font=("Segoe UI", 11, "bold"), insertbackground=TEXT,
                         highlightthickness=1,
                         highlightbackground=ACCENT,
                         highlightcolor="#4fc3f7")
            e.pack(fill="x", ipady=9)
            setattr(self, attr, e)
            # Debounced lookup on every keystroke
            e.bind("<KeyRelease>",
                   lambda evt, a=attr: self._schedule_lookup(a))

        # ── Asset Model (auto-detected) ───────────────────────────────────────
        model_hdr = tk.Frame(frm, bg=CARD)
        model_hdr.pack(fill="x", pady=(9, 2))
        tk.Label(model_hdr, text="Asset Model", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        self.lbl_model_auto = tk.Label(model_hdr, text="", bg=CARD, fg="#4fc3f7",
                                       font=("Segoe UI", 8, "italic"))
        self.lbl_model_auto.pack(side="right")
        self.v_model = tk.Entry(frm, bg=INPUT_BG, fg=TEXT, relief="flat",
                                font=("Segoe UI", 10), insertbackground=TEXT,
                                highlightthickness=1,
                                highlightbackground="#334d66",
                                highlightcolor=ACCENT)
        self.v_model.pack(fill="x", ipady=8)

        # ── Rest of form ──────────────────────────────────────────────────────
        type_row = tk.Frame(frm, bg=CARD)
        type_row.pack(fill="x", pady=(9, 0))
        tk.Frame(type_row, bg=CARD).pack(side="left")  # spacer
        lbl_f = tk.Frame(frm, bg=CARD)
        lbl_f.pack(fill="x", pady=(4, 2))
        tk.Label(lbl_f, text="Asset Type", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        tk.Label(lbl_f, text=" *", bg=CARD, fg=DANGER,
                 font=("Segoe UI", 9, "bold")).pack(side="left")
        self.lbl_autodetect = tk.Label(lbl_f, text="",
                                       bg=CARD, fg="#4fc3f7",
                                       font=("Segoe UI", 8, "italic"))
        self.lbl_autodetect.pack(side="right")
        self.v_type = ttk.Combobox(frm, values=ASSET_TYPES, state="readonly",
                                   font=("Segoe UI", 10))
        self.v_type.set("Laptop")
        self.v_type.pack(fill="x", ipady=4)

        lbl("Assign To", required=True)
        self.v_member = ttk.Combobox(frm, values=TEAM_MEMBERS, state="readonly",
                                     font=("Segoe UI", 10))
        self.v_member.pack(fill="x", ipady=4)

        lbl("Transaction Type", required=True)
        dir_f = tk.Frame(frm, bg=CARD)
        dir_f.pack(fill="x", pady=6)
        self.v_direction = tk.StringVar(value="Out Store")
        for label, val in [("🔴  Out Store", "Out Store"),
                           ("🟢  In Store",  "In Store")]:
            tk.Radiobutton(dir_f, text=label, variable=self.v_direction,
                           value=val, bg=CARD, fg=TEXT, selectcolor=INPUT_BG,
                           font=("Segoe UI", 10),
                           activebackground=CARD,
                           activeforeground=TEXT).pack(side="left", padx=(0, 12))

        lbl("Notes")
        self.v_notes = ent()

        # Bind Enter key on asset_no for barcode scanner (scanner sends Enter after scan)
        self.v_asset_no.bind("<Return>", lambda e: self._on_scan_enter())
        self.v_serial.bind("<Return>",   lambda e: self._on_scan_enter())

        # Buttons
        tk.Frame(panel, bg=CARD, height=8).pack()

        # Single submit button (hidden in bulk mode)
        self.btn_submit = tk.Button(panel, text="  ➤  Submit Transaction",
                  bg=ACCENT, fg="white",
                  font=("Segoe UI", 11, "bold"),
                  relief="flat", cursor="hand2", pady=11,
                  command=self._submit,
                  activebackground=ACCENT2)
        self.btn_submit.pack(fill="x", padx=18, pady=(0, 5))

        tk.Button(panel, text="Clear",
                  bg=CARD2, fg=SUBTEXT, font=("Segoe UI", 9),
                  relief="flat", cursor="hand2", pady=7,
                  command=self._clear_form,
                  activebackground=INPUT_BG).pack(fill="x", padx=18, pady=(0, 6))

        # Bulk mode toggle button
        self.btn_bulk = tk.Button(panel, text="📷  Bulk Scan Mode",
                  bg=CARD2, fg=TEXT, font=("Segoe UI", 9, "bold"),
                  relief="flat", cursor="hand2", pady=7,
                  command=self._toggle_bulk_mode,
                  activebackground=INPUT_BG)
        self.btn_bulk.pack(fill="x", padx=18, pady=(0, 14))

        tk.Button(panel, text="⚙  Server Settings",
                  bg=CARD, fg=SUBTEXT, font=("Segoe UI", 8),
                  relief="flat", cursor="hand2",
                  command=self._open_settings,
                  activebackground=CARD).pack(side="bottom", pady=(0, 2))

    # ── History panel ─────────────────────────────────────────────────────────

    def _build_history(self, parent):
        panel = tk.Frame(parent, bg=CARD)
        panel.pack(side="right", fill="both", expand=True)

        # Header
        hdr = tk.Frame(panel, bg=CARD)
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        tk.Label(hdr, text="Transaction History",
                 bg=CARD, fg=ACCENT,
                 font=("Segoe UI", 12, "bold")).pack(side="left")

        tk.Button(hdr, text="📊  Export to Excel",
                  bg=SUCCESS, fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", cursor="hand2", padx=10, pady=5,
                  command=self._open_export,
                  activebackground="#388e3c").pack(side="right", padx=(6, 0))
        tk.Button(hdr, text="↻  Refresh",
                  bg=CARD2, fg=TEXT, font=("Segoe UI", 9),
                  relief="flat", cursor="hand2", padx=10, pady=5,
                  command=self._refresh_transactions,
                  activebackground=INPUT_BG).pack(side="right")

        # ── Filter / Sort toolbar ─────────────────────────────────────────────
        toolbar = tk.Frame(panel, bg=CARD2, padx=10, pady=7)
        toolbar.pack(fill="x", padx=16, pady=(0, 6))

        def tb_label(text):
            tk.Label(toolbar, text=text, bg=CARD2, fg=SUBTEXT,
                     font=("Segoe UI", 8)).pack(side="left", padx=(6, 2))

        tb_label("Member:")
        self.v_filter_member = ttk.Combobox(
            toolbar, values=["All"] + TEAM_MEMBERS,
            state="readonly", font=("Segoe UI", 8), width=14)
        self.v_filter_member.set("All")
        self.v_filter_member.pack(side="left")
        self.v_filter_member.bind("<<ComboboxSelected>>", lambda _: self._apply_filters())

        tb_label("Status:")
        self.v_filter_status = ttk.Combobox(
            toolbar, values=["All", "Pending", "Confirmed", "Rejected"],
            state="readonly", font=("Segoe UI", 8), width=10)
        self.v_filter_status.set("All")
        self.v_filter_status.pack(side="left")
        self.v_filter_status.bind("<<ComboboxSelected>>", lambda _: self._apply_filters())

        tb_label("Direction:")
        self.v_filter_dir = ttk.Combobox(
            toolbar, values=["All", "Out Store", "In Store"],
            state="readonly", font=("Segoe UI", 8), width=9)
        self.v_filter_dir.set("All")
        self.v_filter_dir.pack(side="left")
        self.v_filter_dir.bind("<<ComboboxSelected>>", lambda _: self._apply_filters())

        tb_label("Type:")
        self.v_filter_type = ttk.Combobox(
            toolbar, values=["All"] + ASSET_TYPES,
            state="readonly", font=("Segoe UI", 8), width=9)
        self.v_filter_type.set("All")
        self.v_filter_type.pack(side="left")
        self.v_filter_type.bind("<<ComboboxSelected>>", lambda _: self._apply_filters())

        # Search
        tk.Label(toolbar, text="  🔍", bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 10)).pack(side="left", padx=(8, 0))
        self.v_search = tk.StringVar()
        self.v_search.trace_add("write", lambda *_: self._apply_filters())
        search_e = tk.Entry(toolbar, textvariable=self.v_search,
                            bg=INPUT_BG, fg=TEXT, relief="flat",
                            font=("Segoe UI", 9), insertbackground=TEXT,
                            highlightthickness=1,
                            highlightbackground="#334d66",
                            highlightcolor=ACCENT, width=18)
        search_e.pack(side="left", ipady=4, padx=(3, 0))
        tk.Button(toolbar, text="✕", bg=CARD2, fg=SUBTEXT,
                  relief="flat", cursor="hand2", font=("Segoe UI", 8),
                  command=lambda: self.v_search.set(""),
                  activebackground=CARD2).pack(side="left", padx=2)

        tk.Button(toolbar, text="Clear All",
                  bg=CARD2, fg=SUBTEXT, font=("Segoe UI", 8),
                  relief="flat", cursor="hand2",
                  command=self._clear_filters,
                  activebackground=INPUT_BG).pack(side="right", padx=6)

        # ── Treeview ──────────────────────────────────────────────────────────
        # Asset No. and Serial No. are first two data columns
        cols   = ("Asset No.", "Serial No.", "Model", "Type",
                  "Member", "Direction", "Status", "Rejection Reason", "Time")
        widths = [90, 90, 130, 72, 110, 70, 105, 160, 118]

        tf = tk.Frame(panel, bg=CARD)
        tf.pack(fill="both", expand=True, padx=16, pady=(0, 4))

        self.tree = ttk.Treeview(tf, columns=cols, show="headings")
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c,
                              command=lambda col=c: self._sort_by(col))
            self.tree.column(c, width=w, anchor="center", minwidth=28)

        sb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.tree.tag_configure("confirmed", foreground=SUCCESS)
        self.tree.tag_configure("rejected",  foreground=DANGER)
        self.tree.tag_configure("pending",   foreground=WARNING)

        self.row_count_lbl = tk.Label(panel, text="",
                                      bg=CARD, fg=SUBTEXT, font=("Segoe UI", 8))
        self.row_count_lbl.pack(anchor="e", padx=18, pady=(2, 8))

    # ── Actions ───────────────────────────────────────────────────────────────

    def _submit(self):
        asset_no = self.v_asset_no.get().strip()
        serial   = self.v_serial.get().strip()
        atype    = self.v_type.get()
        member   = self.v_member.get()
        direc    = self.v_direction.get()

        if not asset_no and not serial:
            messagebox.showerror("Missing ID",
                "Please enter at least the Asset Number or Serial Number.")
            return
        if not member:
            messagebox.showerror("Missing Field", "Please select a Team Member.")
            return

        # ── Validate & correct against lookup sheet ───────────────────────
        # Try asset_no first, then serial_no — accept whichever is found
        found_data = None
        tried = []
        for q in filter(None, [asset_no, serial]):
            if q in tried:
                continue
            tried.append(q)
            try:
                r = api("get", f"{self.server_url}/lookup",
                        params={"q": q}, timeout=4)
                if r.ok:
                    d = r.json()
                    if d.get("found"):
                        found_data = d
                        break
            except requests.exceptions.ConnectionError:
                messagebox.showerror("Connection Error",
                    f"Cannot reach server at:\n{self.server_url}")
                return
            except Exception as e:
                messagebox.showerror("Error", f"Lookup failed: {e}")
                return

        if found_data is None:
            lines = ""
            if asset_no: lines += f"Asset No: {asset_no}\n"
            if serial:   lines += f"Serial No: {serial}\n"
            messagebox.showerror("Asset Not in Lookup",
                f"\u26a0  This asset was not found in the lookup sheet.\n\n"
                f"{lines}\n"
                "Only registered assets can be submitted.")
            return

        # Auto-correct both fields to the canonical values from the lookup sheet
        correct_an   = str(found_data.get("asset_number",  "") or asset_no).strip()
        correct_sn   = str(found_data.get("serial_number", "") or serial).strip()
        correct_type = found_data.get("asset_type", atype)
        asset_no = correct_an
        serial   = correct_sn
        atype    = correct_type if correct_type in ASSET_TYPES else atype
        self.v_asset_no.delete(0, "end")
        self.v_asset_no.insert(0, asset_no)
        self.v_serial.delete(0, "end")
        self.v_serial.insert(0, serial)
        self.v_type.set(atype)

        payload = {
            "asset_number":  asset_no,
            "serial_number": serial,
            "asset_type":    atype,
            "asset_model":   self.v_model.get().strip(),
            "team_member":   member,
            "direction":     direc,
            "notes":         self.v_notes.get().strip(),
            "storekeeper":   "Storekeeper User",
        }
        try:
            r = api("post", f"{self.server_url}/transactions", json=payload)
            if r.status_code == 409:
                messagebox.showerror("Duplicate Pending",
                    r.json().get("error", "This asset already has a pending transaction."))
                return
            if r.status_code == 201:
                id_line = ""
                if asset_no: id_line += f"Asset No : {asset_no}\n"
                if serial:   id_line += f"Serial No: {serial}\n"
                icon = "🔴" if direc == "Out Store" else "🟢"
                messagebox.showinfo("Submitted ✓",
                    f"{icon}  Transaction recorded!\n\n"
                    f"{id_line}"
                    f"Type  : {atype}\n"
                    f"To    : {member}\n"
                    f"Status: {direc}\n\n"
                    f"⏳  Waiting for {member.split()[0]}'s confirmation…")
                self._clear_form()
                self._refresh_transactions()
            else:
                err = r.json().get("error", r.status_code)
                messagebox.showerror("Server Error", err)
        except requests.exceptions.ConnectionError:
            messagebox.showerror("Connection Error",
                f"Cannot reach server at:\n{self.server_url}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _clear_form(self):
        for w in (self.v_asset_no, self.v_serial, self.v_notes):
            w.delete(0, "end")
        self.v_member.set("")
        self.v_direction.set("Out Store")
        self._unlock_type()

    # ── Asset Lookup (auto-fill) ──────────────────────────────────────────────

    def _schedule_lookup(self, source_field):
        """Cancel any pending lookup and schedule a new one in 400 ms."""
        if self._lookup_after_id:
            self.root.after_cancel(self._lookup_after_id)
        q = getattr(self, source_field).get().strip()
        if len(q) < 2:
            return
        self._lookup_after_id = self.root.after(
            400, lambda: self._do_lookup(q, source_field))

    def _do_lookup(self, q, source_field):
        """Hit /lookup and auto-fill sibling field + type combobox."""
        self._lookup_after_id = None
        def fetch():
            try:
                r = api("get", f"{self.server_url}/lookup",
                        params={"q": q}, timeout=4)
                if r.ok:
                    data = r.json()
                    self.root.after(0, lambda: self._apply_lookup(data, source_field))
            except Exception:
                pass  # server offline — silent, don't block the user
        threading.Thread(target=fetch, daemon=True).start()

    def _apply_lookup(self, data, source_field):
        """Called on the main thread after a successful lookup response."""
        if not data.get('found'):
            return
        atype   = data.get('asset_type', '')
        an      = data.get('asset_number', '')
        sn      = data.get('serial_number', '')
        amodel  = data.get('asset_model', '')
        # Fill the sibling field only if it is currently empty
        if source_field == 'v_asset_no' and sn:
            if not self.v_serial.get().strip():
                self.v_serial.delete(0, 'end')
                self.v_serial.insert(0, sn)
        elif source_field == 'v_serial' and an:
            if not self.v_asset_no.get().strip():
                self.v_asset_no.delete(0, 'end')
                self.v_asset_no.insert(0, an)
        # Fill and lock the model field
        self.v_model.config(state='normal')
        self.v_model.delete(0, 'end')
        self.v_model.insert(0, amodel)
        self.v_model.config(state='disabled')
        self.lbl_model_auto.config(text='✦ auto-detected')
        # Set and lock the type combobox
        if atype in ASSET_TYPES:
            self.v_type.set(atype)
            self.v_type.config(state='disabled')
            self.lbl_autodetect.config(text='✦ auto-detected')
            self._lookup_locked = True

    def _unlock_type(self):
        """Restore type and model fields to normal (called by Clear)."""
        self.v_type.config(state='readonly')
        if hasattr(self, 'lbl_autodetect'):
            self.lbl_autodetect.config(text='')
        if hasattr(self, 'v_model'):
            self.v_model.config(state='normal')
            self.v_model.delete(0, 'end')
        if hasattr(self, 'lbl_model_auto'):
            self.lbl_model_auto.config(text='')
        self._lookup_locked = False

    # ── Barcode Scanner & Bulk Mode ───────────────────────────────────────────

    def _on_scan_enter(self):
        """Called when scanner sends Enter after a barcode.
        In bulk mode: immediately add to queue and clear for next scan.
        In single mode: just trigger the lookup (same as KeyRelease)."""
        if self._bulk_mode:
            self._bulk_add_current()
        else:
            # Force immediate lookup (don't wait 400 ms debounce)
            if self._lookup_after_id:
                self.root.after_cancel(self._lookup_after_id)
                self._lookup_after_id = None
            q = self.v_asset_no.get().strip() or self.v_serial.get().strip()
            if q:
                self._do_lookup(q, 'v_asset_no' if self.v_asset_no.get().strip() else 'v_serial')

    def _toggle_bulk_mode(self):
        """Switch between single and bulk scan modes."""
        self._bulk_mode = not self._bulk_mode
        if self._bulk_mode:
            self.btn_bulk.config(
                text="❌  Exit Bulk Mode",
                bg=WARNING, fg="white",
                activebackground="#c47d00")
            self.btn_submit.pack_forget()
            self._open_bulk_panel()
        else:
            self._bulk_mode = False
            self._bulk_queue.clear()
            self.btn_bulk.config(
                text="📷  Bulk Scan Mode",
                bg=CARD2, fg=TEXT,
                activebackground=INPUT_BG)
            self.btn_submit.pack(fill="x", padx=18, pady=(0, 5),
                                 before=self.btn_bulk)
            if hasattr(self, '_bulk_win') and self._bulk_win.winfo_exists():
                self._bulk_win.destroy()
            self._clear_form()

    def _open_bulk_panel(self):
        """Floating panel showing the scan queue."""
        win = tk.Toplevel(self.root)
        win.title("Bulk Scan Mode")
        win.geometry("620x520")
        win.configure(bg=BG)
        win.resizable(True, True)
        win.protocol("WM_DELETE_WINDOW", self._toggle_bulk_mode)
        self._bulk_win = win

        # Header
        hdr = tk.Frame(win, bg=WARNING, height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📷  Bulk Scan Mode",
                 bg=WARNING, fg="white",
                 font=("Segoe UI", 12, "bold")).pack(side="left", padx=16, pady=10)
        tk.Label(hdr,
                 text="Scan barcodes one by one — each adds to the queue below",
                 bg=WARNING, fg="#fff8e1",
                 font=("Segoe UI", 8)).pack(side="right", padx=16)

        # Engineer + Direction row (set once for all)
        cfg_frm = tk.Frame(win, bg=CARD2, padx=16, pady=10)
        cfg_frm.pack(fill="x", padx=14, pady=(10, 4))

        tk.Label(cfg_frm, text="Engineer:", bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self._bulk_member = ttk.Combobox(cfg_frm, values=TEAM_MEMBERS,
                                          state="readonly", font=("Segoe UI", 10), width=20)
        m = self.v_member.get()
        self._bulk_member.set(m if m else TEAM_MEMBERS[0])
        self._bulk_member.grid(row=0, column=1, padx=(0, 20))

        tk.Label(cfg_frm, text="Direction:", bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=2, sticky="w", padx=(0, 8))
        self._bulk_direction = tk.StringVar(value=self.v_direction.get())
        for txt, val in [("🔴  Out Store", "Out Store"),
                         ("🟢  In Store",  "In Store")]:
            tk.Radiobutton(cfg_frm, text=txt, variable=self._bulk_direction,
                           value=val, bg=CARD2, fg=TEXT,
                           selectcolor=INPUT_BG,
                           font=("Segoe UI", 9),
                           activebackground=CARD2,
                           activeforeground=TEXT).grid(row=0, column=3 if val=="Out Store" else 4,
                                                       padx=4)

        # Scan input row
        scan_frm = tk.Frame(win, bg=CARD, padx=14, pady=10)
        scan_frm.pack(fill="x", padx=14, pady=(0, 4))

        tk.Label(scan_frm, text="📷  Scan / type Asset No. or Serial No.:",
                 bg=CARD, fg=ACCENT,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))

        scan_row = tk.Frame(scan_frm, bg=CARD)
        scan_row.pack(fill="x")

        self._bulk_entry = tk.Entry(scan_row, bg=INPUT_BG, fg=TEXT, relief="flat",
                                    font=("Segoe UI", 13, "bold"),
                                    insertbackground=TEXT,
                                    highlightthickness=2,
                                    highlightbackground=ACCENT,
                                    highlightcolor="#4fc3f7")
        self._bulk_entry.pack(side="left", fill="x", expand=True, ipady=10)
        self._bulk_entry.focus_set()
        self._bulk_entry.bind("<Return>", lambda e: self._bulk_add_from_entry())

        tk.Button(scan_row, text="Add ⏎",
                  bg=ACCENT, fg="white",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", padx=12, pady=8,
                  command=self._bulk_add_from_entry,
                  activebackground=ACCENT2).pack(side="left", padx=(8, 0))

        self._bulk_scan_status = tk.Label(scan_frm, text="",
                                           bg=CARD, fg="#4caf50",
                                           font=("Segoe UI", 8, "italic"))
        self._bulk_scan_status.pack(anchor="w", pady=(4, 0))

        # ── Submit All — packed FIRST so it anchors to bottom before table expands
        self._btn_submit_all = tk.Button(
            win,
            text="➤  Submit All  (0 items)",
            bg="#2e7d32", fg="white",
            font=("Segoe UI", 12, "bold"),
            relief="flat", cursor="hand2", pady=12,
            command=self._bulk_submit_all,
            state="disabled",
            activebackground="#1b5e20")
        self._btn_submit_all.pack(fill="x", padx=14, pady=(6, 12), side="bottom")

        # Remove selected — also anchored to bottom above Submit
        tk.Button(win, text="✖  Remove Selected Row  (or press Delete)",
                  bg=CARD2, fg=SUBTEXT, font=("Segoe UI", 8),
                  relief="flat", cursor="hand2", pady=4,
                  command=self._bulk_remove_selected,
                  activebackground=INPUT_BG).pack(fill="x", padx=14,
                                                  pady=(0, 2), side="bottom")

        # ── Queue table header row
        q_hdr = tk.Frame(win, bg=BG)
        q_hdr.pack(fill="x", padx=14, pady=(8, 2))
        self._bulk_count_lbl = tk.Label(q_hdr, text="Queue: 0 items",
                                         bg=BG, fg=SUBTEXT,
                                         font=("Segoe UI", 9, "bold"))
        self._bulk_count_lbl.pack(side="left")
        tk.Button(q_hdr, text="✖  Clear All",
                  bg=BG, fg=DANGER, font=("Segoe UI", 8),
                  relief="flat", cursor="hand2",
                  command=self._bulk_clear_queue,
                  activebackground=BG).pack(side="right")

        # ── Queue treeview — fills remaining space between header and buttons
        tf = tk.Frame(win, bg=BG)
        tf.pack(fill="both", expand=True, padx=14, pady=(0, 4))

        cols = ("Asset No.", "Serial No.", "Type")
        self._bulk_tree = ttk.Treeview(tf, columns=cols, show="headings", height=8)
        for c, w in zip(cols, [160, 180, 100]):
            self._bulk_tree.heading(c, text=c)
            self._bulk_tree.column(c, width=w, anchor="center")
        bulk_sb = ttk.Scrollbar(tf, orient="vertical", command=self._bulk_tree.yview)
        self._bulk_tree.configure(yscrollcommand=bulk_sb.set)
        self._bulk_tree.pack(side="left", fill="both", expand=True)
        bulk_sb.pack(side="right", fill="y")

        # Delete key removes selected row
        self._bulk_tree.bind("<Delete>", lambda e: self._bulk_remove_selected())

        # Restore any already-queued items
        for item in self._bulk_queue:
            self._bulk_tree.insert("", "end",
                values=(item["asset_number"], item["serial_number"], item["asset_type"]))
        self._bulk_refresh_ui()

    # ── Bulk queue helpers ────────────────────────────────────────────────────

    def _bulk_add_from_entry(self):
        """Read the bulk scan entry, look up asset, add to queue."""
        val = self._bulk_entry.get().strip()
        if not val:
            return
        self._bulk_entry.delete(0, "end")
        self._bulk_scan_status.config(text="⏳ Looking up...", fg=SUBTEXT)
        self._bulk_entry.focus_set()

        def fetch():
            try:
                r = api("get", f"{self.server_url}/lookup",
                        params={"q": val}, timeout=4)
                if r.ok:
                    data = r.json()
                    self.root.after(0, lambda: self._bulk_got_lookup(val, data))
                else:
                    self.root.after(0, lambda: self._bulk_got_lookup(val, {"found": False}))
            except Exception:
                self.root.after(0, lambda: self._bulk_got_lookup(val, {"found": False}))
        threading.Thread(target=fetch, daemon=True).start()

    def _bulk_got_lookup(self, scanned_val, data):
        """Apply lookup result and add row to bulk queue."""
        if data.get("found"):
            an     = data.get("asset_number",  "") or scanned_val
            sn     = data.get("serial_number", "")
            atype  = data.get("asset_type",    "Laptop")
            amodel = data.get("asset_model",   "")
            self._bulk_scan_status.config(
                text=f"✔ {scanned_val} → {atype}", fg=SUCCESS)
        else:
            # Not in lookup — reject unknown assets
            self._bulk_scan_status.config(
                text=f"✗ '{scanned_val}' not found in asset list — rejected",
                fg=DANGER)
            return

        # Deduplicate by asset_number+serial_number combo
        key = (an.lower(), sn.lower())
        if any((q["asset_number"].lower(), q["serial_number"].lower()) == key
               for q in self._bulk_queue):
            self._bulk_scan_status.config(
                text=f"⚠ {scanned_val} already in queue — skipped",
                fg=WARNING)
            return

        item = {"asset_number": an, "serial_number": sn, "asset_type": atype, "asset_model": amodel}
        self._bulk_queue.append(item)
        self._bulk_tree.insert("", "end", values=(an, sn, atype))
        self._bulk_refresh_ui()
        if hasattr(self, '_bulk_entry'):
            self._bulk_entry.focus_set()

    def _bulk_add_current(self):
        """Used when scanner fires Enter on the main form in bulk mode.
        Forward to the bulk panel entry if it's open."""
        if hasattr(self, '_bulk_entry') and self._bulk_entry.winfo_exists():
            self._bulk_entry.focus_set()

    def _bulk_remove_selected(self):
        sel = self._bulk_tree.selection()
        if not sel:
            return
        idx = self._bulk_tree.index(sel[0])
        self._bulk_tree.delete(sel[0])
        del self._bulk_queue[idx]
        self._bulk_refresh_ui()

    def _bulk_clear_queue(self):
        for item in self._bulk_tree.get_children():
            self._bulk_tree.delete(item)
        self._bulk_queue.clear()
        self._bulk_refresh_ui()

    def _bulk_refresh_ui(self):
        n = len(self._bulk_queue)
        if hasattr(self, '_bulk_count_lbl'):
            self._bulk_count_lbl.config(text=f"Queue: {n} item{'s' if n != 1 else ''}")
        if hasattr(self, '_btn_submit_all'):
            self._btn_submit_all.config(
                text=f"➤  Submit All  ({n} item{'s' if n != 1 else ''})",
                state="normal" if n > 0 else "disabled",
                bg="#2e7d32" if n > 0 else "#1a2332")

    def _bulk_submit_all(self):
        """Submit every queued item to the server."""
        if not self._bulk_queue:
            return
        member = self._bulk_member.get()
        direc  = self._bulk_direction.get()
        if not member:
            messagebox.showerror("Missing", "Select an engineer first.")
            return

        ok = fail = 0
        for item in self._bulk_queue:
            payload = {
                "asset_number":  item["asset_number"],
                "serial_number": item["serial_number"],
                "asset_type":    item["asset_type"],
                "asset_model":   item.get("asset_model", ""),
                "team_member":   member,
                "direction":     direc,
                "notes":         "",
                "storekeeper":   "Storekeeper User",
            }
            try:
                r = api("post", f"{self.server_url}/transactions", json=payload)
                if r.status_code == 201:
                    ok += 1
                elif r.status_code == 409:
                    fail += 1  # already pending
                else:
                    fail += 1
            except Exception:
                fail += 1

        icon = "🔴" if direc == "Out Store" else "🟢"
        msg  = (f"{icon} {ok} transaction{'s' if ok != 1 else ''} submitted "
                f"to {member.split()[0]}.\n")
        if fail:
            msg += f"\n⚠ {fail} failed — check connection."

        messagebox.showinfo("Bulk Submit Complete", msg)
        self._bulk_clear_queue()
        self._refresh_transactions()

    def _clear_filters(self):
        self.v_filter_member.set("All")
        self.v_filter_status.set("All")
        self.v_filter_dir.set("All")
        self.v_filter_type.set("All")
        self.v_search.set("")

    def _refresh_transactions(self):
        def fetch():
            try:
                r = api("get", f"{self.server_url}/transactions")
                if r.ok:
                    self._all_rows = r.json()
                    self.root.after(0, self._apply_filters)
            except: pass
        threading.Thread(target=fetch, daemon=True).start()

    def _apply_filters(self):
        rows = list(self._all_rows)

        member_f = self.v_filter_member.get() if hasattr(self, 'v_filter_member') else "All"
        status_f = self.v_filter_status.get() if hasattr(self, 'v_filter_status') else "All"
        dir_f    = self.v_filter_dir.get()    if hasattr(self, 'v_filter_dir')    else "All"
        type_f   = self.v_filter_type.get()   if hasattr(self, 'v_filter_type')   else "All"
        query    = self.v_search.get().strip().lower() if hasattr(self, 'v_search') else ""

        if member_f != "All":
            rows = [t for t in rows if t['team_member'] == member_f]
        if dir_f != "All":
            rows = [t for t in rows if t['direction'] == dir_f]
        if type_f != "All":
            rows = [t for t in rows if t.get('asset_type') == type_f]
        if status_f == "Pending":
            rows = [t for t in rows if not t['confirmed'] and not t['rejected']]
        elif status_f == "Confirmed":
            rows = [t for t in rows if t['confirmed']]
        elif status_f == "Rejected":
            rows = [t for t in rows if t['rejected']]
        if query:
            rows = [t for t in rows if any(
                query in str(t.get(f, "")).lower()
                for f in ('asset_number', 'serial_number', 'asset_type',
                          'team_member', 'notes', 'direction', 'rejection_reason')
            )]

        for item in self.tree.get_children():
            self.tree.delete(item)

        for t in rows:
            if   t['confirmed']: status, tag = "✅ Confirmed", "confirmed"
            elif t['rejected']:  status, tag = "❌ Rejected",  "rejected"
            else:                status, tag = "⏳ Pending",   "pending"

            ts = t.get('timestamp', '')
            try: ts = datetime.fromisoformat(ts).strftime("%d/%m/%y  %H:%M")
            except: pass

            direc  = "🔴 Out" if t['direction'] == "Out Store" else "🟢 In"
            fname  = t['team_member'].split()[0]
            reason = t.get('rejection_reason', '') or ''

            self.tree.insert("", "end", values=(
                t.get('asset_number',  ''),
                t.get('serial_number', ''),
                t.get('asset_model',   ''),
                t.get('asset_type',    ''),
                fname, direc, status, reason, ts
            ), tags=(tag,))

        count = len(rows)
        total = len(self._all_rows)
        self.row_count_lbl.config(
            text=f"{count} record{'s' if count != 1 else ''}"
                 + (f"  (filtered from {total})" if count != total else "")
        )

    def _sort_by(self, col):
        col_map = {
            "Asset No.": "asset_number", "Serial No.": "serial_number",
            "Model": "asset_model", "Type": "asset_type",
            "Member": "team_member", "Direction": "direction",
            "Status": "confirmed", "Rejection Reason": "rejection_reason",
            "Time": "timestamp"
        }
        key = col_map.get(col, col)
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = False
        self._all_rows.sort(
            key=lambda x: str(x.get(key, "")),
            reverse=self._sort_rev
        )
        self._apply_filters()

    def _open_export(self):
        ExportDialog(self.root, self.server_url)

    def _open_settings(self):
        """Open settings — requires admin password."""
        auth = tk.Toplevel(self.root)
        auth.title("Admin Login")
        auth.geometry("360x205")
        auth.configure(bg=BG)
        auth.resizable(False, False)
        auth.grab_set()
        auth.attributes("-topmost", True)

        tk.Label(auth, text="🔐  Admin Authentication",
                 bg=BG, fg=TEXT,
                 font=("Segoe UI", 12, "bold")).pack(pady=(16, 4), padx=20, anchor="w")
        tk.Label(auth, text="Admin password required to change settings.",
                 bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(padx=20, anchor="w")

        frm = tk.Frame(auth, bg=CARD2, padx=20, pady=14)
        frm.pack(fill="x", padx=20, pady=10)
        tk.Label(frm, text="Password", bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 4))
        pw_var = tk.StringVar()
        pw_e = tk.Entry(frm, textvariable=pw_var, show="●",
                        bg=INPUT_BG, fg=TEXT, relief="flat",
                        font=("Segoe UI", 11), insertbackground=TEXT)
        pw_e.pack(fill="x", ipady=8)
        pw_e.focus_set()

        err = tk.Label(auth, text="", bg=BG, fg=DANGER, font=("Segoe UI", 9))
        err.pack(pady=(0, 4))

        def verify():
            try:
                r = api("post", f"{self.server_url}/admin/verify",
                        json={"password": pw_var.get()})
                if r.ok and r.json().get("valid"):
                    auth.destroy()
                    self._show_settings_window()
                else:
                    err.config(text="❌  Wrong password.")
            except Exception:
                err.config(text="⚠  Cannot reach server.")

        pw_e.bind("<Return>", lambda _: verify())
        tk.Button(auth, text="Login",
                  bg=ACCENT, fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", pady=9,
                  command=verify,
                  activebackground=ACCENT2).pack(fill="x", padx=20)

    def _show_settings_window(self):
        """Settings window shown after admin authentication."""
        win = tk.Toplevel(self.root)
        win.title("Server Settings")
        win.geometry("400x180")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="⚙  Server Settings", bg=BG, fg=TEXT,
                 font=("Segoe UI", 12, "bold")).pack(pady=14, padx=20, anchor="w")

        frm = tk.Frame(win, bg=CARD2, padx=20, pady=14)
        frm.pack(fill="x", padx=20)
        tk.Label(frm, text="Server URL", bg=CARD2, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 4))
        url_e = tk.Entry(frm, bg=INPUT_BG, fg=TEXT, relief="flat",
                         font=("Segoe UI", 10), insertbackground=TEXT)
        url_e.insert(0, self.server_url)
        url_e.pack(fill="x", ipady=8)

        def save():
            new_url = url_e.get().strip().rstrip("/")
            self.server_url = new_url
            self.cfg["server_url"] = new_url
            save_cfg(self.cfg)
            win.destroy()
            self._check_server()

        tk.Button(win, text="Save", bg=ACCENT, fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                  command=save, pady=9).pack(fill="x", padx=20, pady=12)

    def _check_server(self):
        if self._server_check_timer:
            self.root.after_cancel(self._server_check_timer)
            self._server_check_timer = None
        def _do():
            try:
                r = api("get", f"{self.server_url}/health")
                if r.ok:
                    total = r.json().get('total', 0)
                    self.status_lbl.config(
                        text=f"● Online  |  {total} records",
                        fg="#90ee90")
                else:
                    self.status_lbl.config(text="● Server Error", fg=WARNING)
            except Exception:
                self.status_lbl.config(
                    text="● Server Offline — check Admin Settings", fg=DANGER)
        threading.Thread(target=_do, daemon=True).start()
        self._server_check_timer = self.root.after(20_000, self._check_server)


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    StorekeeperApp(root)
    root.mainloop()
