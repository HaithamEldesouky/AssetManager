"""
Asset Manager — Server
- HTTP / HTTPS (internal domain network)
- SQLite database
- Admin password protected settings
- Asset lookup from Excel file (asset_lookup.xlsx)
- Minimises to system tray
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import sqlite3, json, os, sys, socket, threading, time, hashlib, secrets, base64, io, ctypes
from datetime import datetime, timedelta

# pywin32 — optional; required only for Windows Service mode
try:
    import win32serviceutil, win32service, win32event, servicemanager
    HAS_WIN32SVC = True
except ImportError:
    HAS_WIN32SVC = False

try:
    import pystray
    from PIL import Image, ImageDraw
    HAS_TRAY = True
except ImportError:
    HAS_TRAY = False

def _make_server_tray_icon():
    img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
    d   = ImageDraw.Draw(img)
    d.rectangle([8, 24, 56, 56], fill=(0, 120, 212))
    d.rectangle([20, 12, 44, 28], fill=(0, 120, 212))
    d.line([(32, 12), (32, 56)], fill=(255, 255, 255, 180), width=2)
    d.line([(8, 36), (56, 36)],  fill=(255, 255, 255, 180), width=2)
    d.ellipse([42, 44, 58, 60], fill=(76, 175, 80))  # green dot = running
    return img

APP_VERSION = "1.6.0"   # bump when releasing a new build; clients compare against this

app = Flask(__name__)
CORS(app)

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH         = os.path.join(BASE_DIR, "assets.db")
ADMIN_CFG_PATH  = os.path.join(BASE_DIR, "admin_config.json")
SERVER_CFG_PATH = os.path.join(BASE_DIR, "server_config.json")
LOOKUP_PATH     = os.path.join(BASE_DIR, "asset_lookup.xlsx")
SSL_CERT_PATH   = os.path.join(BASE_DIR, "ssl_cert.pem")
SSL_KEY_PATH    = os.path.join(BASE_DIR, "ssl_key.pem")

def _load_server_cfg():
    default = {"host": "0.0.0.0", "port": 8081}
    if os.path.exists(SERVER_CFG_PATH):
        try:
            with open(SERVER_CFG_PATH) as f:
                d = json.load(f)
            default.update(d)
        except Exception:
            pass
    return default

_SERVER_CFG = _load_server_cfg()
HOST = _SERVER_CFG["host"]
PORT = _SERVER_CFG["port"]


def _port_in_use(port):
    """Return True if something is already listening on 127.0.0.1:port."""
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(1)
            return s.connect_ex(("127.0.0.1", port)) == 0
    except Exception:
        return False


def _acquire_single_instance_mutex():
    """
    Create a Global Named Mutex so only one interactive GUI process runs.
    Returns True if this process is the first (mutex created fresh).
    Returns False if another GUI instance already holds it.
    Only meaningful on Windows; always returns True on other platforms.
    """
    if sys.platform != "win32":
        return True
    try:
        handle = ctypes.windll.kernel32.CreateMutexW(None, True,
                                                      "Global\\AssetManagerServer_GUI")
        err = ctypes.windll.kernel32.GetLastError()
        # ERROR_ALREADY_EXISTS == 183
        return err != 183
    except Exception:
        return True  # can't check → allow


ASSET_TYPES = ["Laptop", "Desktop", "Mobile", "Tablet", "Screen", "UPS", "Server",
               "Cisco Phone", "Printer", "Scanner", "Switch"]

# ─── Embedded default lookup (asset_lookup.xlsx shipped with the app) ─────────
DEFAULT_LOOKUP_B64 = (
    "UEsDBBQAAAAIAPyKzlxGx01IlwAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE2PTQvCMBBE/0ro3aS16EFiQdSj6Ml7TDc2kGSXZIX476WCH7cZhvdg9CUjQWYPRdQYUtk2EzNtlCp2gmiKRIJUY3CYo+EiMd8VOuctHNA+IiRWy7ZdK6gMaYRxQV9hM+gdUfDWsMc0nLzNWNCxOFYLQewxkmF/CyCUOBMketYgetnJlVb/4Gy5Qi5z7mX3Hj9dq9+B4QVQSwMEFAAAAAgA/IrOXBPssUnvAAAAKwIAABEAAABkb2NQcm9wcy9jb3JlLnhtbM2SQWrDMBBFrxK0t0eyUy+E401KVikUGmjpTkiTRNSyhDTFzu2L1cShtAfodubz/huYVgepfcTn6ANGsphWk+uHJHXYsDNRkABJn9GpVPqAw+T6o49OUSp9PEFQ+kOdECrOG3BIyihSMAOLsBBZ1xotdURFPl7xRi/48Bn7DDMasEeHAyUQpQDWzY3hMvUt3AEzjDC69D1AsxDz9E9s3gC7Jqdkl9Q4juVY51zFuYC3p/1LPrewQyI1aGSrKVlJl4Abdmt+rbePhx3rKl41BW8KsT6Itaxq+dC8z64//O7Czht7tP/Y+CbYtfDrL7ovUEsDBBQAAAAIAPyKzlyZXJwjCQYAAJwnAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbO1a31PbOBB+56/Q6Gbu7Ro7jkNCMR2cH+Wu0DKQ600fN45iq8iSR1KA/Pc3skmwHMehnVDaO/KAY1nft/utV7uWw/G7+5ShWyIVFTzA7hsHvzs5OIYjnZCUoPuUcXUEAU60zo5aLRUlJAX1RmSE36dsLmQKWr0RMm7NJNxRHqes1XacbisFyjHikJIAf5rPaUTQxFDikwOEVvwjRlLCtTJj+WjE5LUxQSxkjnmYMbtxV2f5uVqqAZPoFliA7yifibsJudcYMVB6wGSAnfyDW2uOlkVyDEdM76Is0Y3zj01XIsg9bNt0Mp6u+dxxp384rHrTtrxpgI9Go8HIrVovwyGKCK8KKlN0xj03rHhQAa1pGjwZOL7TqaXZ9MbbTtMPw9Dv19F4GzSd7TQ9p9s5bdfRdDZo/IbYhKeDQbeOxt+g6W6nGR/2u51amjXoGI4SRvnNdhKTtdVEsyDHcDQX7KyZpec4Tq+S/TbKjKyX3XohzgXXO1ZiCl+FHAuuLesMNOVILzMyh4gEeADpVFJ49CCfRaA0pXItUtuvGbeQiiTNdID/yoDj0tzff7sfj9vDt/nR896i4otjBjwn7DwcD4vjwCuOp+O3TUbOgMcVI2F/6Brs4LDj5EZOB6PcSOi7/i4yVSHzw15osJ2x7+3C6gq264e53cNhIbLbdUbm2D8ddhq5TiVMy1wTmhKFPpI7dCVS4I1ukKn8TugkAWpBIREpNCFGOrEQH5fAGgEhse/WZ0n5rBHxfvHV0nOdyIWmTYgPSWohLoRgoZDN2j8YN8raFzze4ZdclAFXALeNbg0quTVaZAlJN1aejUmIJeWSAdcQE040MtfEDSFN+C+UWvfngkZSKDHX6AtFIdDmQE7oVNejz2gKDJaNvk8SsCJ68RmFgjUaHJJbGwI8BtZohDDrLryHhYa0WRWkrAw5B500Crleysi6cUpL4DFhAo1mRKlG8Ce5tCR9AEZ3ZNYFW6Y2RGp60wg5ByHKkKG4GSSQZs26KE/KoD/VjRAM0KXQzf4Jew2bc8Eo8N0Z9ZkS/Z3F6W8aJ/XJaK4spN1CN3qf6YeUP6kfMjqVVRWv/fCn6oenkjbXhWoX3An4j/a+ISz4JeHJa+t7bX2vre9nan07K9I3Njy7uRXbyNUW8XHXmO7aNM4pY9d6yci5svukEozOxpSxx9FiPOdb72ezZMBKrhWe1GCP4SiWkA8iKfQ/VCfXCWQkwO7andU8ZfmyHkWZUAF2rOlNTlXnFa+5KNfFJN9+DWXzgb4Qs2KeV3lfZQld2a242zL+bpXgGdP7kuEdvpQMt2Dckw7Xf6IOfw86ipFKmpmHQ8oR8DjAbrddqEMqAkZmJk0rSb5K558vx1UCM/KQ5O7Toup6+84O86Jrfzr63kvp2EeWl4V0nirEf5E0d3aled5papqGoeW1nYRxdBfgvt/2MYogC/CcgcYoSrNZgJVpsMBiHuBI2+Hb1oSeHvxK6LdEtBJ4p27a1rBvaXc5bSaVHoJKCuJ8VjW6jNeEqu13zC153li1nluF13N/VRXFWU2Gk/mcRLo2y0uXKqaLK3X1Xiw0kdfJ7A5N2UJewSzApjw4GM2o0gFur05kgE1O5Gd2Z6mvTNXfLWoKWPHLCcsSeOirve31pqDbXBFr/6t3oUby43AlRs8VO+8Hxq6hVr/G7mVj91A7CCfebCMQEaREAjLFIcBC6kTEErKERmMpuK6TKIVGDLQJAGLmF3oTGXJbaZwrfwr+DbOMxom+ojGSNA6wTiQhl/oh3t9m1X3o3zW2V0Y2KuRmLEyEsprwTMktYRNTzLvmNmGUrJrTZt218FsStjJs19ZpPP7f7kWL1feDNj+WhMLyvmQ07eFKD2L9l1K754f5oj/vFtL2n/FhPgOdIPMnwBGVEXt8vbOeYp7XJ+KKRBqtX3wgHeA/ik0aMmW++DYNsFsMbqxwY+JX2QE/pmTP+ZVfj5RyzXtqru1DyDPkml+TajXr+2mZZsbq+kW+OV298zRDZmDjP9vME9D0K4n0kMxhwbTKPTBPTPdawmD1vzfnSrdODtYMJwf/AlBLAwQUAAAACAD8is5ckJV9xeQDAABNGgAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbJ2ZXW+bPBhA/wrifgWbfE4EqfM2vZPWKVr3ce0Qp0EFzGu7y/bvJ9IWx9Pz8KDcgcWxj6MoR8T5SZtHe1TKRb+burWb+Ohc9zZJbHlUjbQ3ulPt76Y+aNNIZ2+0eUhsZ5Tcn6GmTniaLpJGVm1c5OexrSly/eTqqlVbE9mnppHmzztV69MmZvHrwNfq4ej6gaTIO/mg7pX73m1NUuTJMMu+alRrK91GRh028S17K7JZD5yf+FGpk724jvqt7LR+7G8+7Tdx2hupWpWun0KWrvqlhKrrfqY4sv+/TBoPa/bg5fXr7B/Pm9+aaCetErr+We3dcROv4mivDvKpdl/16T/1sqH5IPheOlnkRp8i02+0yMv+ol/bbeKq7T+ge2fiIq9skbvi1lrloi9PzU6ZPHFFnvTjSfnCvcO4e2UqWeOgGF/w259OhVRi9Gnw5oM3x9a/vdt+/vAmTVMGaaPYlzeeBFGBoZ9l53Q3Ip0N0hkpzSHpbJI0hIrsWunZID0jpTNIejZJGkLF7Frp+SA9J6VnkPR8kjSECgx9r+zjuPVisF6Q1nPIejHJGkLF4mrr5WC9JK0XkPVykjWEiuXV1qvBekVaLyHr1SRrCBUYeqd3VT32q7cepNek9AqSXk+ShlCxvlaapb4xKam9BhuTTvKGWIGytPhFHNHKva7OUlCcTREHWYGy3+SuVm5M3NeRkXlkYB5xLhAH+4iytLgvJCMTycBE4lwgDjYSZWlxX0lGZpKBmcS5QBzsJM6WRql2TNyXkpGpZGAqcS4QB1uJs6S4jyUja8nAWuJcIA7mEmdJcd9LRgaTgcHEuUAcLCbKft/ej1n7XjIymAwMJs4F1mAxUZaw9sFkZDEZWEycC6zBZKLsuDX3weRkMBkYTJwLrMFg4qwyv/59lQvFfTA5GUwOBhPnLsVBVuAsKX7xOkkGk8Pvk5OCCbICZ0lxH0xOBpODwcS5QBwMJsqKypY62h51O/oS76vJyWpysJo4F9iD1UTZifY+nZxMJwfTiXOBPZhOlJ1o7/vJyX5ysJ84F9iD/UTZralaN/6d9wHlZEA5GFCcC8zBgKLsBHMfUU5GlIMRxbnAHIwoyk4w9yHlZEg5GFKcC8zBkOJsKdt21DzzMc3ImHIwpjgXmIMxxVna3Nc0I2uagTXFuUtzkBU4S5v7nGZkTjMwpzgXmIM5xdlT5crjmPjFP7RkTjP4L9pJOQVZgbOkuC9pRpY0A0uKc4E4WFKcRcWfjzOezyL6g5Y7aR6q1ka1OrhNnN4s53Fkns8unm+c7s4HNTvtnG7Ol0cl98r0D8zj6KC1e73pj06GE6TiL1BLAwQUAAAACAD8is5cfPOj3FkCAAD2CQAADQAAAHhsL3N0eWxlcy54bWzdVtuK2zAQ/RXjD1glMTVRkfXQQKDQloXdh74qsZwIdHEleUn69WUkJ2l2NQtL32oTPJqjM3ehsBDPWj4dpYzVyWgbuvoY4/iZkLA/SiPCgxulPRk9OG9EDA/OH0gYvRR9AJLRZLVYtMQIZWvO7GS2JoZq7yYbu3pRE84GZ2+aZZ0VnFlhZPUidFdvhFY7r9JeYZQ+Z/UKFHunna/iURoJZMJZ+J3hZV5BlLMdo6zzoCTZQ/oEzgal9TWAFQSgtOZsFDFKb7dK68RJyjdQNcvP51F29cGL83L1KTtJW9MncLZzvpf+Ls+s4kzLIRLOvDoc4RvdSACM0RnCWa/EwVmRYrgwZiFwtpdaP0GLfg53tk9DlWv9tYcyV5DqRVRaz2I2kxdg/29r2fY/m61G9eLilylGZ9P61+SifPRyUKe0Pg33/q+uUyB33q/aCoajq3/AzOmbiWo3KR2VnVdH1ffSvsktcBbFTst7+4u66uUgJh2fr2BX3+TvsleTodddj5DWvOsmf4MeLtvrZAbOlO3lSfabeekPuyRW/rDr6sX8AOE1sk1PGcE4GSsjgGF+sAgwTmZhfv6nfNZoPhnDYlsXkTXKWaOczCohm/RifsocSiktZ0pp07QtVtHNphjBBqtb28KvbA2LDRiYH/D0sVrj3cYn5P05wHr63oRgmeKTiGWK1xqQct2AQWm525gfYGBdwGYH/Jf9wEyVOU0DXcViw04wjlCKITCL5RltW6Q6Lbzl/mCnpGkoLSOAlSNoGgyB04gjWAQQA4Y0TboHX91H5HJPkds/Pf4HUEsDBBQAAAAIAPyKzlyXirscwAAAABMCAAALAAAAX3JlbHMvLnJlbHOd0ktqAzEMgOGrDN53lKbQRchk1U12peQCqq15MLYlZJW6tw/JJpnQF9mLn09C2zeKaBPnMk5SmppiLp0bzWQDUPxICUvLQrmm2LMmtNKyDiDoZxwI1qvVM+h1w+22183m8CX0nyL3/eTphf1HomzfhG8mXHNAHcg6VyN8ss7vzHNbU3TNPnRO9+HRwZ0W+XE7SGQY0BA8Kz2IspDaROXCCexflaWcJxag9d2gv49D1SgHCr+TUGQhejqJYPEDuyNQSwMEFAAAAAgA/IrOXNv5o+o8AQAAKQIAAA8AAAB4bC93b3JrYm9vay54bWyNUdFKxDAQ/JWQD7j0DhU82oIo6sGh4onv23R7XS7JlmTrqV8vaTk88MWnzcyGYWa2PHI8NMwH9eldSJXuRYa1Mcn26CEteMDw6V3H0YOkBce9SUNEaFOPKN6ZVVFcGQ8UdF2etF6iOQcsaIU4mLrMxDvhMf3uM1QflKghR/JV6entUCtPgTx9Y1vpQqvU8/GRI31zEHA7G9m5Si/nxTtGIfuH3mWTb9CkiRFoXkGIK31VFFp1FJNMPyZ9sEIf+AbNjEbhe3KC8Q4EHyKPA4V9ljF1ac5iTD2c5lziOv6nRu46snjHdvQYZO4xossGQ+ppSFoF8Fjpm5RQ1Jb5MA45FqJs2jmigOBZYXFNbaXjpp1dnqy12FHA9gk8JlOXFpx9iSqPSWd1cbm81qobnbsFZ5/DlmHSzxqnK9U/UEsDBBQAAAAIAPyKzlwkHpuirAAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO1kb0Kg0AQhF/luAdw1UCKoFZp0gZf4NDVE++P2w3Rtw/RQoUUaayWmeKbYad4olE8eEd6CCQmaxyVUjOHGwA1Gq2ixAd0kzWdj1YxJT72EFQzqh4hT9MrxD1DVsWeKeo54D9E33VDg3ffvCw6/gGGt48jaUSWolaxRy4lTGazCZaTJZM1UjzaUsZHm0k4rRDxbJC2Nqs+xOcnxrNGi1v6Ilfz+ILLtwMchq4+UEsDBBQAAAAIAPyKzlxlkHmSGQEAAM8DAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbK2Tz07DMAzGX6XKdWo8OHBA6y7AFXbgBULirtGSOIq90b09avdHAo1qaFziQ+zvZ/tLFu/7jFz1MSRuVCeSHwHYdhgNa8qY+hhaKtEIaypryMZuzBrhfj5/AEtJMEktg4ZaLp6xNdsg1UsvmNhTalTBwKp6OiQOrEaZnIO3Rjwl2CX3g1IfCbpgGHO485lnfQwKLhKGm98Bx7q3HZbiHVYrU+TVRGwU9AFY9gFZT0tc6JHa1lt0ZLcRk2jOBY3jDlFi0AfR2TRZOox4OO9u5o8yU0BHdlUoM1gq+HfcyZKhus6FMhbx0yOeiSbnm+fDwW2H7kp2H+CTymb0g2EMt+/4u8dn/Sv6+CDa/PcTG6KOxqcTH8Z/vPwCUEsBAhQAFAAAAAgA/IrOXEbHTUiXAAAAzQAAABAAAAAAAAAAAAAAAIABAAAAAGRvY1Byb3BzL2FwcC54bWxQSwECFAAUAAAACAD8is5cE+yxSe8AAAArAgAAEQAAAAAAAAAAAAAAgAHFAAAAZG9jUHJvcHMvY29yZS54bWxQSwECFAAUAAAACAD8is5cmVycIwkGAACcJwAAEwAAAAAAAAAAAAAAgAHjAQAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQIUABQAAAAIAPyKzlyQlX3F5AMAAE0aAAAYAAAAAAAAAAAAAAC2gR0IAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECFAAUAAAACAD8is5cfPOj3FkCAAD2CQAADQAAAAAAAAAAAAAAgAE3DAAAeGwvc3R5bGVzLnhtbFBLAQIUABQAAAAIAPyKzlyXirscwAAAABMCAAALAAAAAAAAAAAAAACAAbsOAABfcmVscy8ucmVsc1BLAQIUABQAAAAIAPyKzlzb+aPqPAEAACkCAAAPAAAAAAAAAAAAAACAAaQPAAB4bC93b3JrYm9vay54bWxQSwECFAAUAAAACAD8is5cJB6boqwAAAD4AQAAGgAAAAAAAAAAAAAAgAENEQAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECFAAUAAAACAD8is5cZZB5khkBAADPAwAAEwAAAAAAAAAAAAAAgAHxEQAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLBQYAAAAACQAJAD4CAAA7EwAAAAA="
)

# ─── Helpers ──────────────────────────────────────────────────────────────────

def sha256(text: str) -> str:
    return hashlib.sha256(text.encode()).hexdigest()

# ─── Admin config ─────────────────────────────────────────────────────────────

def load_admin_cfg():
    if os.path.exists(ADMIN_CFG_PATH):
        with open(ADMIN_CFG_PATH) as f:
            return json.load(f)
    cfg = {"password_hash": sha256("admin"), "first_run": True}
    save_admin_cfg(cfg)
    return cfg

def save_admin_cfg(cfg):
    with open(ADMIN_CFG_PATH, "w") as f:
        json.dump(cfg, f, indent=2)

ADMIN_CFG = load_admin_cfg()

# ─── Database ─────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        asset_type       TEXT    NOT NULL,
        serial_number    TEXT    DEFAULT '',
        asset_number     TEXT    DEFAULT '',
        team_member      TEXT    NOT NULL,
        direction        TEXT    NOT NULL,
        timestamp        TEXT    NOT NULL,
        confirmed        INTEGER DEFAULT 0,
        rejected         INTEGER DEFAULT 0,
        confirmed_at     TEXT    DEFAULT '',
        rejection_reason TEXT    DEFAULT '',
        notes            TEXT    DEFAULT '',
        storekeeper      TEXT    DEFAULT 'Storekeeper User'
    )''')
    cols = [row[1] for row in c.execute("PRAGMA table_info(transactions)").fetchall()]
    if 'rejection_reason' not in cols:
        c.execute("ALTER TABLE transactions ADD COLUMN rejection_reason TEXT DEFAULT ''")
    if 'asset_number' not in cols:
        c.execute("ALTER TABLE transactions ADD COLUMN asset_number TEXT DEFAULT ''")
    if 'asset_model' not in cols:
        c.execute("ALTER TABLE transactions ADD COLUMN asset_model TEXT DEFAULT ''")
    c.execute("""CREATE TABLE IF NOT EXISTS team_members (
        name         TEXT PRIMARY KEY,
        auto_approve INTEGER DEFAULT 0
    )""")
    if c.execute("SELECT COUNT(*) FROM team_members").fetchone()[0] == 0:
        for nm in ["Engineer 1", "Engineer 2", "Engineer 3", "Engineer 4"]:
            c.execute("INSERT OR IGNORE INTO team_members(name, auto_approve) VALUES (?,0)", (nm,))
    conn.commit()
    conn.close()


_ARCHIVE_COLS = ["id", "asset_type", "serial_number", "asset_number", "asset_model",
                 "team_member", "direction", "timestamp", "confirmed", "rejected",
                 "confirmed_at", "rejection_reason", "notes", "storekeeper"]

def _export_records_xlsx(path, ids=None):
    import openpyxl
    conn = get_db()
    if ids:
        qmarks = ",".join("?" * len(ids))
        rows = conn.execute(
            f"SELECT {','.join(_ARCHIVE_COLS)} FROM transactions "
            f"WHERE id IN ({qmarks}) ORDER BY id", list(ids)).fetchall()
    else:
        rows = conn.execute(f"SELECT {','.join(_ARCHIVE_COLS)} FROM transactions ORDER BY id").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Records Archive"
    ws.append(_ARCHIVE_COLS)
    for r in rows:
        ws.append([r[c] for c in _ARCHIVE_COLS])
    wb.save(path)
    return len(rows)

def _import_records_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if not data:
        return 0
    header = [str(h).strip() if h else "" for h in data[0]]
    idx = {h: i for i, h in enumerate(header)}
    insert_cols = [c for c in _ARCHIVE_COLS if c != "id"]
    conn = get_db()
    n = 0
    for row in data[1:]:
        if not row:
            continue
        vals = []
        for c in insert_cols:
            v = row[idx[c]] if c in idx and idx[c] < len(row) else None
            vals.append('' if v is None else v)
        conn.execute(
            f"INSERT INTO transactions ({','.join(insert_cols)}) "
            f"VALUES ({','.join(['?'] * len(insert_cols))})", vals)
        n += 1
    conn.commit()
    conn.close()
    return n

def simple_password_prompt(parent, title, message):
    from tkinter import simpledialog
    return simpledialog.askstring(title, message + "\n\nAdmin password:", show="*", parent=parent)

# ─── Lookup helpers ───────────────────────────────────────────────────────────

def _ensure_sample_lookup():
    """Write the embedded real asset_lookup.xlsx if no file exists yet."""
    if os.path.exists(LOOKUP_PATH):
        return
    try:
        raw = base64.b64decode(DEFAULT_LOOKUP_B64)
        with open(LOOKUP_PATH, 'wb') as f:
            f.write(raw)
    except Exception:
        pass  # fallback: leave missing, server will still run


def _load_lookup_table():
    """
    Load asset_lookup.xlsx into a dict keyed by both asset_number and serial_number
    (both lowercased for case-insensitive lookup).
    Returns dict: lower(value) -> {"asset_type": ..., "asset_number": ..., "serial_number": ...}
    """
    _ensure_sample_lookup()
    if not os.path.exists(LOOKUP_PATH):
        return {}
    try:
        import openpyxl
        wb  = openpyxl.load_workbook(LOOKUP_PATH, read_only=True, data_only=True)
        ws  = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return {}

    table = {}
    for row in rows:
        if not row or len(row) < 3:
            continue
        asset_no  = str(row[0]).strip() if row[0] else ""
        serial_no = str(row[1]).strip() if row[1] else ""
        atype     = str(row[2]).strip() if row[2] else ""
        amodel    = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        if amodel.upper() in ("N/A", "N\\A", "NA", "N.A", "-"):
            amodel = ""
        if atype not in ASSET_TYPES:
            continue
        entry = {"asset_type": atype, "asset_number": asset_no, "serial_number": serial_no, "asset_model": amodel}
        if asset_no:
            table[asset_no.lower()] = entry
        if serial_no:
            table[serial_no.lower()] = entry
    return table


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    conn  = get_db()
    count = conn.execute('SELECT COUNT(*) FROM transactions').fetchone()[0]
    conn.close()
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat(), "total": count})

@app.route('/version')
def get_version():
    """Latest published version + download link (if the installer is present next to
    AssetServer.exe). Clients compare this against their own APP_VERSION and prompt
    the user to update when behind. Drop a new AssetManager_Setup.exe beside the
    server exe and bump APP_VERSION to push an update."""
    installer = os.path.join(BASE_DIR, "AssetManager_Setup.exe")
    has_installer = os.path.exists(installer)
    return jsonify({
        "latest": APP_VERSION,
        "has_installer": has_installer,
        "download_path": "/download/installer" if has_installer else "",
    })

@app.route('/download/installer')
def download_installer():
    installer = os.path.join(BASE_DIR, "AssetManager_Setup.exe")
    if not os.path.exists(installer):
        return jsonify({"error": "installer not available"}), 404
    return send_file(installer, as_attachment=True,
                     download_name="AssetManager_Setup.exe")

@app.route('/transactions', methods=['GET'])
def get_transactions():
    member = request.args.get('member', '')
    conn   = get_db()
    if member:
        rows = conn.execute(
            'SELECT * FROM transactions WHERE team_member=? ORDER BY timestamp DESC', (member,)
        ).fetchall()
    else:
        rows = conn.execute('SELECT * FROM transactions ORDER BY timestamp DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/transactions', methods=['POST'])
def create_transaction():
    data = request.json
    if not data.get('team_member') or not data.get('direction'):
        return jsonify({"error": "Missing required fields"}), 400
    if not data.get('serial_number') and not data.get('asset_number'):
        return jsonify({"error": "Provide at least Serial Number or Asset Number"}), 400

    asset_no  = (data.get('asset_number',  '') or '').strip()
    serial_no = (data.get('serial_number', '') or '').strip()

    conn = get_db()

    # Block duplicate pending transactions for the same asset
    conditions, params = [], []
    if asset_no:
        conditions.append("(asset_number=? AND asset_number != '')")
        params.append(asset_no)
    if serial_no:
        conditions.append("(serial_number=? AND serial_number != '')")
        params.append(serial_no)
    if conditions:
        existing = conn.execute(
            f"SELECT id FROM transactions WHERE ({' OR '.join(conditions)}) AND confirmed=0 AND rejected=0",
            params
        ).fetchone()
        if existing:
            conn.close()
            return jsonify({"error": "This asset already has a pending transaction. Confirm or reject it first."}), 409

    auto_row = conn.execute(
        "SELECT auto_approve FROM team_members WHERE name=?", (data['team_member'],)
    ).fetchone()
    auto_approve = bool(auto_row and auto_row['auto_approve'])
    now_iso = datetime.now().isoformat()

    cur  = conn.execute(
        '''INSERT INTO transactions
           (asset_type, serial_number, asset_number, asset_model,
            team_member, direction, timestamp, notes, storekeeper,
            confirmed, confirmed_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (
            data.get('asset_type',   'Laptop'),
            serial_no,
            asset_no,
            (data.get('asset_model', '') or '').strip(),
            data['team_member'],
            data['direction'],
            now_iso,
            data.get('notes', ''),
            data.get('storekeeper', 'Storekeeper User'),
            1 if auto_approve else 0,
            now_iso if auto_approve else '',
        )
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"id": new_id, "status": "auto_approved" if auto_approve else "created",
                    "auto_approved": auto_approve}), 201

@app.route('/pending/<path:member>')
def get_pending(member):
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM transactions WHERE team_member=? AND confirmed=0 AND rejected=0 ORDER BY timestamp',
        (member,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/confirm/<int:trans_id>', methods=['POST'])
def confirm_transaction(trans_id):
    conn = get_db()
    conn.execute('UPDATE transactions SET confirmed=1, confirmed_at=? WHERE id=?',
                 (datetime.now().isoformat(), trans_id))
    conn.commit()
    conn.close()
    return jsonify({"status": "confirmed"})

@app.route('/reject/<int:trans_id>', methods=['POST'])
def reject_transaction(trans_id):
    data   = request.json or {}
    reason = data.get('reason', '')
    conn   = get_db()
    conn.execute(
        'UPDATE transactions SET rejected=1, confirmed_at=?, rejection_reason=? WHERE id=?',
        (datetime.now().isoformat(), reason, trans_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"status": "rejected"})

@app.route('/stats')
def get_stats():
    conn      = get_db()
    total     = conn.execute('SELECT COUNT(*) FROM transactions').fetchone()[0]
    pending   = conn.execute('SELECT COUNT(*) FROM transactions WHERE confirmed=0 AND rejected=0').fetchone()[0]
    confirmed = conn.execute('SELECT COUNT(*) FROM transactions WHERE confirmed=1').fetchone()[0]
    out_now   = conn.execute("SELECT COUNT(*) FROM transactions WHERE direction='Out Store' AND confirmed=1").fetchone()[0]
    conn.close()
    return jsonify({"total": total, "pending": pending, "confirmed": confirmed, "out_now": out_now})

@app.route('/export')
def export_all():
    conn = get_db()
    rows = conn.execute('SELECT * FROM transactions ORDER BY timestamp DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/members', methods=['GET'])
def list_members():
    conn = get_db()
    rows = conn.execute('SELECT name, auto_approve FROM team_members ORDER BY name').fetchall()
    conn.close()
    return jsonify([{"name": r["name"], "auto_approve": bool(r["auto_approve"])} for r in rows])

@app.route('/members', methods=['POST'])
def add_member():
    data = request.json or {}
    if sha256(data.get('password', '')) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong admin password"}), 401
    name = (data.get('name', '') or '').strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    auto = 1 if data.get('auto_approve') else 0
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO team_members(name, auto_approve) VALUES (?,?)", (name, auto))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "name": name, "auto_approve": bool(auto)})

@app.route('/members/delete', methods=['POST'])
def delete_member():
    """Admin-only: remove an engineer from the assignee list. Existing transaction
    history is unaffected (team_member is stored as text on each record)."""
    data = request.json or {}
    if sha256(data.get('password', '')) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong admin password"}), 401
    name = (data.get('name', '') or '').strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    conn = get_db()
    cur = conn.execute("DELETE FROM team_members WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "deleted": cur.rowcount, "name": name})

@app.route('/transactions/delete', methods=['POST'])
def delete_transactions():
    data = request.json or {}
    if sha256(data.get('password', '')) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong admin password"}), 401
    ids = data.get('ids', [])
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "No record ids provided"}), 400
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid ids"}), 400
    conn = get_db()
    qmarks = ",".join("?" * len(ids))
    cur = conn.execute(f"DELETE FROM transactions WHERE id IN ({qmarks})", ids)
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"status": "ok", "deleted": deleted})

# ─── Lookup routes ────────────────────────────────────────────────────────────

@app.route('/lookup')
def lookup_asset():
    """
    GET /lookup?q=<asset_number_or_serial>
    Returns {"found": true, "asset_type": "...", "asset_number": "...", "serial_number": "..."}
    or      {"found": false}
    """
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"found": False, "error": "No query provided"}), 400
    table = _load_lookup_table()
    entry = table.get(q.lower())
    if entry:
        return jsonify({"found": True, **entry})
    return jsonify({"found": False})

@app.route('/lookup/file')
def lookup_file():
    """GET /lookup/file — Download the current lookup Excel file."""
    _ensure_sample_lookup()
    if not os.path.exists(LOOKUP_PATH):
        return jsonify({"error": "No lookup file on server"}), 404
    return send_file(
        LOOKUP_PATH,
        as_attachment=True,
        download_name="asset_lookup.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/lookup/upload', methods=['POST'])
def lookup_upload():
    """
    POST /lookup/upload
    Body: {"password": "...", "file_data": "<base64-encoded xlsx>"}
    Replaces asset_lookup.xlsx on the server.
    """
    data = request.json or {}
    if sha256(data.get('password', '')) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong admin password"}), 401
    file_b64 = data.get('file_data', '')
    if not file_b64:
        return jsonify({"error": "No file data provided"}), 400
    try:
        raw = base64.b64decode(file_b64)
        # Quick validation — check it's a valid xlsx
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        wb.close()
        with open(LOOKUP_PATH, 'wb') as f:
            f.write(raw)
        # Count rows loaded
        table = _load_lookup_table()
        unique = {(e['asset_number'], e['serial_number']) for e in table.values()}
        return jsonify({"status": "ok", "assets_loaded": len(unique)})
    except Exception as e:
        return jsonify({"error": f"Invalid Excel file: {e}"}), 400

# ─── Admin routes ─────────────────────────────────────────────────────────────

@app.route('/admin/verify', methods=['POST'])
def admin_verify():
    data = request.json or {}
    if sha256(data.get('password', '')) == ADMIN_CFG['password_hash']:
        return jsonify({"valid": True})
    return jsonify({"valid": False}), 401

@app.route('/admin/change_password', methods=['POST'])
def admin_change_password():
    data    = request.json or {}
    current = data.get('current_password', '')
    new_pw  = data.get('new_password', '')
    if sha256(current) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong current password"}), 401
    if len(new_pw) < 4:
        return jsonify({"error": "Password too short (min 4 chars)"}), 400
    ADMIN_CFG['password_hash'] = sha256(new_pw)
    ADMIN_CFG['first_run']     = False
    save_admin_cfg(ADMIN_CFG)
    return jsonify({"status": "ok"})

@app.route('/admin/server_config', methods=['GET'])
def get_server_config():
    data = request.json or {}
    if sha256(data.get('password', '')) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong admin password"}), 401
    return jsonify({"host": HOST, "port": PORT})

@app.route('/admin/server_config', methods=['POST'])
def set_server_config():
    data = request.json or {}
    if sha256(data.get('password', '')) != ADMIN_CFG['password_hash']:
        return jsonify({"error": "Wrong admin password"}), 401
    new_host = str(data.get('host', '0.0.0.0')).strip()
    try:
        new_port = int(data.get('port', 80))
        if not (1 <= new_port <= 65535):
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid port (1–65535)"}), 400
    with open(SERVER_CFG_PATH, 'w') as f:
        json.dump({"host": new_host, "port": new_port}, f, indent=2)
    return jsonify({"status": "ok", "note": "Restart the server to apply changes."})

@app.route('/cert')
def serve_cert():
    if os.path.exists(SSL_CERT_PATH):
        return send_file(SSL_CERT_PATH, mimetype='application/x-pem-file',
                         download_name='ssl_cert.pem', as_attachment=False)
    return jsonify({"error": "No certificate installed on this server"}), 404

# ─── Server GUI ───────────────────────────────────────────────────────────────

def _load_lookup_raw():
    """Return list of (asset_number, serial_number, asset_type, asset_model) from xlsx."""
    _ensure_sample_lookup()
    if not os.path.exists(LOOKUP_PATH):
        return []
    try:
        import openpyxl
        wb   = openpyxl.load_workbook(LOOKUP_PATH, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        result = []
        for row in rows:
            if not row or len(row) < 3:
                continue
            an = str(row[0]).strip() if row[0] else ""
            sn = str(row[1]).strip() if row[1] else ""
            at = str(row[2]).strip() if row[2] else ""
            am = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            if am.upper() in ("N/A", "N\\A", "NA", "N.A", "-"):
                am = ""
            if not an and not sn:
                continue
            result.append((an, sn, at, am))
        return result
    except Exception:
        return []


def _save_lookup_raw(rows):
    """Write list of (asset_number, serial_number, asset_type) to xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Lookup"
    blue_fill = PatternFill("solid", fgColor="0078D4")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
    headers   = ["Asset Number", "Serial Number", "Asset Type", "Asset Model"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = blue_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 26
    for ri, row_data in enumerate(rows, 2):
        an = row_data[0] if len(row_data) > 0 else ""
        sn = row_data[1] if len(row_data) > 1 else ""
        at = row_data[2] if len(row_data) > 2 else ""
        am = row_data[3] if len(row_data) > 3 else ""
        ws.cell(row=ri, column=1, value=an)
        ws.cell(row=ri, column=2, value=sn)
        ws.cell(row=ri, column=3, value=at)
        ws.cell(row=ri, column=4, value=am)
    wb.save(LOOKUP_PATH)


def run_gui():
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.title("Asset Manager — Server")
    root.geometry("530x760")
    root.minsize(530, 720)
    root.configure(bg="#1e2a3a")
    root.resizable(False, True)

    _tray_icon = [None]
    _minimised_to_tray = [False]

    def _show_window():
        root.after(0, lambda: (root.deiconify(), root.lift(), root.focus_force()))
        _minimised_to_tray[0] = False

    def _hide_window():
        root.after(0, root.withdraw)
        _minimised_to_tray[0] = True

    def _quit_server(*_):
        if messagebox.askyesno("Quit", "Stop the Asset Manager server?"):
            if _tray_icon[0]:
                threading.Thread(target=_tray_icon[0].stop, daemon=True).start()
            os._exit(0)

    def _on_close():
        if HAS_TRAY:
            _hide_window()
            if not _minimised_to_tray[0]:
                # First time — show balloon hint (tray icon title acts as tooltip)
                pass
        else:
            _quit_server()

    root.protocol("WM_DELETE_WINDOW", _on_close)

    if HAS_TRAY:
        def _build_server_tray():
            def _toggle(*_):
                if _minimised_to_tray[0]:
                    _show_window()
                else:
                    _hide_window()
            menu = pystray.Menu(
                pystray.MenuItem("📦  Asset Manager  —  Server", None, enabled=False),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem("🖥  Show / Hide",  _toggle, default=True),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem("❌  Exit Server",  _quit_server),
            )
            icon = pystray.Icon("AssetServer", _make_server_tray_icon(),
                                "Asset Manager — Server  ●  RUNNING", menu)
            _tray_icon[0] = icon
            icon.run()
        threading.Thread(target=_build_server_tray, daemon=True).start()

    hdr = tk.Frame(root, bg="#0078d4", height=55)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(hdr, text="📦  Asset Manager  —  Server",
             bg="#0078d4", fg="white",
             font=("Segoe UI", 14, "bold")).pack(side="left", padx=18, pady=14)

    card = tk.Frame(root, bg="#243447", padx=24, pady=16)
    card.pack(fill="x", padx=18, pady=14)

    hostname = socket.gethostname()
    try:    local_ip = socket.gethostbyname(hostname)
    except: local_ip = "127.0.0.1"

    _ssl_active = (_SERVER_CFG.get('ssl_enabled') and
                   os.path.exists(SSL_CERT_PATH) and os.path.exists(SSL_KEY_PATH))
    _proto      = "https" if _ssl_active else "http"
    server_url  = f"{_proto}://asset-server:{PORT}"

    def info_row(label, value, color="#ffffff"):
        f = tk.Frame(card, bg="#243447")
        f.pack(fill="x", pady=3)
        tk.Label(f, text=label, bg="#243447", fg="#8a9bb0",
                 font=("Segoe UI", 9), width=18, anchor="w").pack(side="left")
        tk.Label(f, text=value, bg="#243447", fg=color,
                 font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left")

    _enc_text  = "HTTPS  (SSL encrypted)" if _ssl_active else "HTTP  (internal network)"
    _enc_color = "#4caf50"

    info_row("Status:",      "● RUNNING",  "#4caf50")
    info_row("Encryption:",  _enc_text,    _enc_color)
    info_row("Host:",        hostname)
    info_row("IP Address:",  local_ip,     "#0078d4")
    info_row("Port:",        str(PORT))
    info_row("Server URL:",  server_url,   "#f0a500")
    info_row("Database:",    DB_PATH)

    def copy_url():
        root.clipboard_clear()
        root.clipboard_append(server_url)
        btn.config(text="✓  Copied!")
        root.after(2500, lambda: btn.config(text="📋  Copy Server URL"))

    btn = tk.Button(root, text="📋  Copy Server URL",
                    bg="#0078d4", fg="white",
                    font=("Segoe UI", 10, "bold"),
                    relief="flat", cursor="hand2", pady=10,
                    command=copy_url, activebackground="#106ebe")
    btn.pack(fill="x", padx=18, pady=(0, 6))

    # ─── Change Admin Password ────────────────────────────────────────────
    def change_admin_password():
        import tkinter.messagebox as mb
        dlg = tk.Toplevel(root)
        dlg.title("Change Admin Password")
        dlg.geometry("400x400")
        dlg.configure(bg="#1e2a3a")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(root)

        hdr = tk.Frame(dlg, bg="#0078d4", height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🔑  Change Admin Password",
                 bg="#0078d4", fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=12)

        frm = tk.Frame(dlg, bg="#243447", padx=20, pady=14)
        frm.pack(fill="x", padx=18, pady=14)

        def pw_field(label):
            tk.Label(frm, text=label, bg="#243447", fg="#8a9bb0",
                     font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 2))
            e = tk.Entry(frm, show="●", bg="#2a3f55", fg="#ffffff",
                         relief="flat", font=("Segoe UI", 10),
                         insertbackground="#ffffff")
            e.pack(fill="x", ipady=7)
            return e

        cur_e = pw_field("Current Password")
        new_e = pw_field("New Password  (min 4 characters)")
        cnf_e = pw_field("Confirm New Password")

        err_lbl = tk.Label(dlg, text="", bg="#1e2a3a", fg="#f44336",
                           font=("Segoe UI", 9))
        err_lbl.pack(pady=(2, 0))

        def do_change():
            cur = cur_e.get()
            new = new_e.get()
            cnf = cnf_e.get()
            if not cur or not new or not cnf:
                err_lbl.config(text="✗  All fields are required.")
                return
            if new != cnf:
                err_lbl.config(text="✗  New password and confirmation do not match.")
                return
            if len(new) < 4:
                err_lbl.config(text="✗  New password must be at least 4 characters.")
                return
            if sha256(cur) != ADMIN_CFG['password_hash']:
                err_lbl.config(text="✗  Current password is incorrect.")
                return
            ADMIN_CFG['password_hash'] = sha256(new)
            ADMIN_CFG['first_run']     = False
            save_admin_cfg(ADMIN_CFG)
            dlg.destroy()
            mb.showinfo("Done", "✓  Admin password changed successfully.")

        tk.Button(dlg, text="Change Password",
                  bg="#0078d4", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", pady=9,
                  command=do_change,
                  activebackground="#106ebe").pack(fill="x", padx=18, pady=(4, 12))
        cur_e.focus_set()

    tk.Button(root, text="🔑  Change Admin Password",
              bg="#f0a500", fg="white",
              font=("Segoe UI", 9, "bold"),
              relief="flat", cursor="hand2", pady=8,
              command=change_admin_password,
              activebackground="#c47d00").pack(fill="x", padx=18, pady=(0, 4))

    def change_host_port():
        import tkinter.messagebox as mb
        dlg = tk.Toplevel(root)
        dlg.title("Change Host / Port")
        dlg.geometry("400x340")
        dlg.configure(bg="#1e2a3a")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(root)

        hdr2 = tk.Frame(dlg, bg="#0078d4", height=44)
        hdr2.pack(fill="x")
        hdr2.pack_propagate(False)
        tk.Label(hdr2, text="🌐  Change Host / Port",
                 bg="#0078d4", fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=12)

        frm2 = tk.Frame(dlg, bg="#243447", padx=20, pady=14)
        frm2.pack(fill="x", padx=18, pady=14)

        def cfg_field(label, default_val):
            tk.Label(frm2, text=label, bg="#243447", fg="#8a9bb0",
                     font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 2))
            e = tk.Entry(frm2, bg="#2a3f55", fg="#ffffff", relief="flat",
                         font=("Segoe UI", 10), insertbackground="#ffffff")
            e.insert(0, str(default_val))
            e.pack(fill="x", ipady=7)
            return e

        host_e = cfg_field("Bind Host  (0.0.0.0 = all interfaces)", HOST)
        port_e = cfg_field("Port", PORT)

        tk.Label(dlg, text="⚠  Restart the server after saving for changes to take effect.",
                 bg="#1e2a3a", fg="#f0a500", font=("Segoe UI", 8),
                 wraplength=360).pack(pady=(0, 2))
        err2 = tk.Label(dlg, text="", bg="#1e2a3a", fg="#f44336", font=("Segoe UI", 9))
        err2.pack()

        def do_save_cfg():
            h = host_e.get().strip()
            p = port_e.get().strip()
            if not h:
                err2.config(text="✗  Host cannot be empty.")
                return
            try:
                p_int = int(p)
                if not (1 <= p_int <= 65535):
                    raise ValueError
            except ValueError:
                err2.config(text="✗  Port must be a number between 1 and 65535.")
                return
            with open(SERVER_CFG_PATH, 'w') as f:
                json.dump({"host": h, "port": p_int}, f, indent=2)
            dlg.destroy()
            mb.showinfo("Saved", "✓  Server config saved.\nRestart the server to apply.")

        tk.Button(dlg, text="Save  (restart to apply)",
                  bg="#0078d4", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", pady=9,
                  command=do_save_cfg,
                  activebackground="#106ebe").pack(fill="x", padx=18, pady=(0, 12))
        host_e.focus_set()

    tk.Button(root, text="🌐  Change Host / Port",
              bg="#2a5e8a", fg="white",
              font=("Segoe UI", 9, "bold"),
              relief="flat", cursor="hand2", pady=8,
              command=change_host_port,
              activebackground="#1a4060").pack(fill="x", padx=18, pady=(0, 4))

    def import_archive():
        import tkinter.messagebox as mb
        import tkinter.filedialog as fd
        pw = simple_password_prompt(root, "Import Archive",
                                    "Restore records from an archive file (.xlsx).\n"
                                    "Records are appended to the current database.")
        if pw is None:
            return
        if sha256(pw) != ADMIN_CFG['password_hash']:
            mb.showerror("Import Archive", "Incorrect admin password.")
            return
        path = fd.askopenfilename(parent=root, title="Open records archive",
                                  filetypes=[("Excel archive", "*.xlsx")])
        if not path:
            return
        try:
            n = _import_records_xlsx(path)
            mb.showinfo("Import Archive", f"Imported {n} record(s) from the archive.")
        except Exception as ex:
            mb.showerror("Import Archive", f"Import failed: {ex}")

    tk.Button(root, text="📥  Import Archive",
              bg="#2a5e8a", fg="white", font=("Segoe UI", 9, "bold"),
              relief="flat", cursor="hand2", pady=8,
              command=import_archive, activebackground="#1a4060").pack(fill="x", padx=18, pady=(0, 4))

    # ─── Manage / Delete Records (with archive-on-delete) ─────────────────────
    def manage_records():
        import tkinter.messagebox as mb
        from tkinter import ttk
        win = tk.Toplevel(root)
        win.title("Manage Records")
        win.geometry("780x500")
        win.configure(bg="#1e2a3a")
        win.transient(root)
        tk.Label(win, text="🗂  Manage Records  —  select rows (or Select All), then Delete",
                 bg="#1e2a3a", fg="#e8edf2", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=14, pady=(12, 6))
        cols = ("id", "asset_number", "serial_number", "type", "member", "direction", "status", "date")
        headers = {cc: cc.replace("_", " ").title() for cc in cols}
        tree = ttk.Treeview(win, columns=cols, show="headings", selectmode="extended", height=15)
        widths = {"id": 45, "asset_number": 95, "serial_number": 95, "type": 80,
                  "member": 130, "direction": 75, "status": 80, "date": 120}
        sort_state = {"col": None, "rev": False}
        def _sort(col):
            rev = (not sort_state["rev"]) if sort_state["col"] == col else False
            sort_state["col"], sort_state["rev"] = col, rev
            def key(v):
                if col == "id":
                    try: return (0, int(v))
                    except Exception: return (1, str(v).lower())
                if col == "date":
                    try: return (0, datetime.strptime(v, "%d/%m/%Y %H:%M"))
                    except Exception: return (1, str(v))
                return (0, str(v).lower())
            items = sorted(((tree.set(k, col), k) for k in tree.get_children("")),
                           key=lambda t: key(t[0]), reverse=rev)
            for pos, (_, k) in enumerate(items):
                tree.move(k, "", pos)
            for cc in cols:
                arrow = ("  ▲" if not rev else "  ▼") if cc == col else ""
                tree.heading(cc, text=headers[cc] + arrow)
        for cc in cols:
            tree.heading(cc, text=headers[cc], command=lambda cc=cc: _sort(cc))
            tree.column(cc, width=widths[cc], anchor="w")
        tree.pack(fill="both", expand=True, padx=14, pady=4)
        def load_rows():
            tree.delete(*tree.get_children())
            conn = get_db()
            rows = conn.execute("SELECT * FROM transactions ORDER BY id DESC").fetchall()
            conn.close()
            for r in rows:
                status = "Confirmed" if r["confirmed"] else ("Rejected" if r["rejected"] else "Pending")
                ds = r["timestamp"]
                try: ds = datetime.fromisoformat(ds).strftime("%d/%m/%Y %H:%M")
                except Exception: pass
                tree.insert("", "end", values=(r["id"], r["asset_number"], r["serial_number"],
                            r["asset_type"], r["team_member"], r["direction"], status, ds))
        load_rows()

        def select_all():
            tree.selection_set(tree.get_children())

        def _confirm_and_delete(ids, all_selected):
            """Password + optional archive, then delete the given ids. Used for any
            count — a subset or every row — so the export option is always offered."""
            import tkinter.filedialog as fd
            dlg = tk.Toplevel(win)
            dlg.title("Delete Records")
            dlg.geometry("420x300")
            dlg.configure(bg="#1e2a3a")
            dlg.resizable(False, False)
            dlg.grab_set()
            dlg.transient(win)
            hdr = tk.Frame(dlg, bg="#c0392b", height=44)
            hdr.pack(fill="x"); hdr.pack_propagate(False)
            scope = "ALL records" if all_selected else f"{len(ids)} selected record(s)"
            tk.Label(hdr, text="🗑  Delete Records", bg="#c0392b", fg="white",
                     font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=12)
            frm = tk.Frame(dlg, bg="#243447", padx=20, pady=14)
            frm.pack(fill="x", padx=18, pady=14)
            tk.Label(frm, text=f"⚠  This will permanently delete {scope}.",
                     bg="#243447", fg="#f0a500", font=("Segoe UI", 9),
                     wraplength=340).pack(anchor="w", pady=(0, 10))
            tk.Label(frm, text="Admin Password", bg="#243447", fg="#8a9bb0",
                     font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 2))
            pw_e = tk.Entry(frm, show="●", bg="#2a3f55", fg="#ffffff", relief="flat",
                            font=("Segoe UI", 10), insertbackground="#ffffff")
            pw_e.pack(fill="x", ipady=7)
            archive_var = tk.IntVar(value=1)
            tk.Checkbutton(frm, text="Export these records to an archive file first",
                           variable=archive_var, bg="#243447", fg="#e8edf2",
                           selectcolor="#2a3f55", activebackground="#243447",
                           activeforeground="#e8edf2", font=("Segoe UI", 9)).pack(anchor="w", pady=(10, 0))
            err_lbl = tk.Label(dlg, text="", bg="#1e2a3a", fg="#f44336", font=("Segoe UI", 9))
            err_lbl.pack(pady=(2, 0))
            def do_delete():
                pw = pw_e.get()
                if not pw:
                    err_lbl.config(text="✗  Password is required."); return
                if sha256(pw) != ADMIN_CFG['password_hash']:
                    err_lbl.config(text="✗  Incorrect password."); return
                save_path = None
                if archive_var.get():
                    default_name = "AssetManager_Archive_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
                    save_path = fd.asksaveasfilename(
                        parent=dlg, title="Save records archive",
                        defaultextension=".xlsx", initialfile=default_name,
                        filetypes=[("Excel archive", "*.xlsx")])
                    if not save_path:
                        err_lbl.config(text="✗  Archive cancelled — nothing deleted."); return
                    try:
                        _export_records_xlsx(save_path, ids=ids)
                    except Exception as ex:
                        err_lbl.config(text=f"✗  Archive failed: {ex}"); return
                if not mb.askyesno("Confirm Delete",
                                   f"Delete {scope} permanently?\nThis cannot be undone.",
                                   icon="warning", parent=dlg):
                    return
                try:
                    conn = get_db()
                    qmarks = ",".join("?" * len(ids))
                    conn.execute(f"DELETE FROM transactions WHERE id IN ({qmarks})", ids)
                    conn.commit(); conn.close()
                    dlg.destroy()
                    load_rows()
                    msg = f"✓  Deleted {len(ids)} record(s)."
                    if save_path:
                        msg += f"\n\nArchived to:\n{save_path}"
                    mb.showinfo("Done", msg)
                except Exception as ex:
                    err_lbl.config(text=f"✗  Error: {ex}")
            tk.Button(dlg, text="🗑  Delete", bg="#c0392b", fg="white",
                      font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2", pady=9,
                      command=do_delete, activebackground="#922b21").pack(fill="x", padx=18, pady=(4, 12))
            pw_e.focus_set()

        def delete_selected():
            sel = tree.selection()
            if not sel:
                mb.showinfo("Manage Records", "Select one or more rows first (or use Select All).")
                return
            all_rows = tree.get_children()
            ids = [int(tree.item(s)["values"][0]) for s in sel]
            _confirm_and_delete(ids, all_selected=(len(sel) == len(all_rows) and len(all_rows) > 0))

        btnf = tk.Frame(win, bg="#1e2a3a")
        btnf.pack(fill="x", padx=14, pady=(4, 12))
        tk.Button(btnf, text="🗑  Delete Selected", bg="#c0392b", fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", pady=7,
                  command=delete_selected, activebackground="#922b21").pack(side="left")
        tk.Button(btnf, text="☑  Select All", bg="#2a5e8a", fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", pady=7,
                  command=select_all, activebackground="#1a4060").pack(side="left", padx=8)
        tk.Button(btnf, text="↻  Refresh", bg="#2a5e8a", fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", pady=7,
                  command=load_rows, activebackground="#1a4060").pack(side="left")

    tk.Button(root, text="🗂  Manage / Delete Records",
              bg="#2a5e8a", fg="white", font=("Segoe UI", 9, "bold"),
              relief="flat", cursor="hand2", pady=8,
              command=manage_records, activebackground="#1a4060").pack(fill="x", padx=18, pady=(0, 4))

    # ─── SSL Certificate ──────────────────────────────────────────────────────
    def _cert_to_pem(cert_path, key_path=None, pfx_password=None):
        """Convert cert (and key) to PEM bytes. Returns (cert_pem, key_pem)."""
        from cryptography import x509
        from cryptography.hazmat.primitives import serialization
        from cryptography.hazmat.primitives.serialization import pkcs12
        from cryptography.hazmat.primitives.serialization import (
            load_pem_private_key, load_der_private_key)

        ext = os.path.splitext(cert_path)[1].lower()

        if ext in ('.pfx', '.p12'):
            with open(cert_path, 'rb') as f:
                data = f.read()
            # Try provided password, then empty-string fallback (Windows exports
            # sometimes require b"" rather than None even when no password is set)
            pwd = pfx_password.encode('utf-8') if pfx_password else None
            last_err = None
            result = None
            for attempt_pwd in ([pwd] if pwd is not None else [None, b""]):
                try:
                    result = pkcs12.load_key_and_certificates(data, attempt_pwd)
                    break
                except Exception as e:
                    last_err = e
            if result is None:
                # Last-resort fallback via pyOpenSSL (handles older/legacy PFX formats)
                try:
                    from OpenSSL import crypto as _ossl
                    _pfx = _ossl.load_pkcs12(data, pwd if pwd is not None else b"")
                    cert_pem = _ossl.dump_certificate(_ossl.FILETYPE_PEM, _pfx.get_certificate())
                    key_pem  = _ossl.dump_privatekey(_ossl.FILETYPE_PEM, _pfx.get_privatekey())
                    return cert_pem, key_pem
                except Exception as oe:
                    pass
                raise ValueError(
                    "Cannot open PFX file.\n"
                    "• If it has a password, enter it in the PFX Password field.\n"
                    f"• Technical detail: {last_err}"
                )
            priv_key, cert_obj, _ = result
            if cert_obj is None:
                raise ValueError("No certificate found in PFX file.")
            if priv_key is None:
                raise ValueError("No private key found in PFX file — "
                                 "make sure the PFX includes the private key.")
            cert_pem = cert_obj.public_bytes(serialization.Encoding.PEM)
            key_pem  = priv_key.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.TraditionalOpenSSL,
                encryption_algorithm=serialization.NoEncryption()
            )
            return cert_pem, key_pem

        # .crt / .cer / .pem — certificate only; key file required
        with open(cert_path, 'rb') as f:
            cert_data = f.read()
        try:
            x509.load_pem_x509_certificate(cert_data)
            cert_pem = cert_data
        except Exception:
            try:
                cert_obj = x509.load_der_x509_certificate(cert_data)
                cert_pem = cert_obj.public_bytes(serialization.Encoding.PEM)
            except Exception:
                raise ValueError("Cannot parse certificate — unsupported format.")

        if not key_path or not os.path.exists(key_path):
            raise ValueError("Private key file is required for .crt / .cer / .pem certificates.")
        with open(key_path, 'rb') as f:
            key_data = f.read()
        try:
            load_pem_private_key(key_data, password=None)
            key_pem = key_data
        except Exception:
            try:
                key_obj = load_der_private_key(key_data, password=None)
                key_pem = key_obj.private_bytes(
                    encoding=serialization.Encoding.PEM,
                    format=serialization.PrivateFormat.TraditionalOpenSSL,
                    encryption_algorithm=serialization.NoEncryption()
                )
            except Exception:
                raise ValueError("Cannot parse key file — unsupported format.")
        return cert_pem, key_pem

    def upload_ssl_cert():
        import tkinter.messagebox as mb
        import tkinter.filedialog as fd

        dlg = tk.Toplevel(root)
        dlg.title("SSL Certificate")
        dlg.geometry("440x440")
        dlg.configure(bg="#1e2a3a")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(root)

        hdr = tk.Frame(dlg, bg="#1a7a4a", height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🔒  SSL Certificate",
                 bg="#1a7a4a", fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=12)

        _ssl_on  = _SERVER_CFG.get('ssl_enabled', False)
        _cert_ok = os.path.exists(SSL_CERT_PATH)
        _status  = "Active (HTTPS)" if (_ssl_on and _cert_ok) else "Inactive (HTTP)"
        _s_color = "#4caf50" if (_ssl_on and _cert_ok) else "#f44336"

        sf = tk.Frame(dlg, bg="#243447", padx=16, pady=10)
        sf.pack(fill="x", padx=18, pady=(10, 0))
        tk.Label(sf, text=f"Status:  {_status}", bg="#243447", fg=_s_color,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        if _cert_ok:
            tk.Label(sf, text=SSL_CERT_PATH, bg="#243447", fg="#8a9bb0",
                     font=("Segoe UI", 8), wraplength=380).pack(anchor="w", pady=(2, 0))

        frm = tk.Frame(dlg, bg="#1e2a3a", padx=18)
        frm.pack(fill="x", pady=(6, 0))

        CERT_TYPES = [
            ("Certificate files", "*.pfx *.p12 *.crt *.cer *.pem"),
            ("PFX / PKCS#12",    "*.pfx *.p12"),
            ("CRT / CER",        "*.crt *.cer"),
            ("PEM",              "*.pem"),
            ("All files",        "*.*"),
        ]
        KEY_TYPES = [
            ("Key files", "*.pem *.key"),
            ("All files", "*.*"),
        ]

        # ── Cert row ──
        tk.Label(frm, text="Certificate file  (.pfx  .crt  .cer  .pem)",
                 bg="#1e2a3a", fg="#8a9bb0", font=("Segoe UI", 9)).pack(anchor="w", pady=(8, 2))
        cert_row = tk.Frame(frm, bg="#1e2a3a")
        cert_row.pack(fill="x")
        cert_var = tk.StringVar()
        cert_e = tk.Entry(cert_row, textvariable=cert_var, bg="#2a3f55", fg="#ffffff",
                          relief="flat", font=("Segoe UI", 9), insertbackground="#ffffff")
        cert_e.pack(side="left", fill="x", expand=True, ipady=6)

        # ── Key row (hidden for PFX) ──
        key_lbl_var = tk.StringVar(value="Private key file  (.pem  .key)")
        key_lbl = tk.Label(frm, textvariable=key_lbl_var,
                           bg="#1e2a3a", fg="#8a9bb0", font=("Segoe UI", 9))
        key_lbl.pack(anchor="w", pady=(8, 2))
        key_row = tk.Frame(frm, bg="#1e2a3a")
        key_row.pack(fill="x")
        key_var = tk.StringVar(value=SSL_KEY_PATH if os.path.exists(SSL_KEY_PATH) else "")
        key_e = tk.Entry(key_row, textvariable=key_var, bg="#2a3f55", fg="#ffffff",
                         relief="flat", font=("Segoe UI", 9), insertbackground="#ffffff")
        key_e.pack(side="left", fill="x", expand=True, ipady=6)
        key_btn = tk.Button(key_row, text="Browse", bg="#2a5e8a", fg="white",
                            relief="flat", cursor="hand2", font=("Segoe UI", 9),
                            command=lambda: key_var.set(
                                fd.askopenfilename(title="Select Key File",
                                                   filetypes=KEY_TYPES) or key_var.get()))
        key_btn.pack(side="left", padx=(4, 0), ipadx=8, ipady=6)

        # ── PFX password row (shown only for PFX) ──
        pfx_lbl = tk.Label(frm, text="PFX Password  (leave blank if none)",
                           bg="#1e2a3a", fg="#8a9bb0", font=("Segoe UI", 9))
        pfx_pwd_var = tk.StringVar()
        pfx_e = tk.Entry(frm, textvariable=pfx_pwd_var, show="●",
                         bg="#2a3f55", fg="#ffffff", relief="flat",
                         font=("Segoe UI", 9), insertbackground="#ffffff")

        def _on_cert_change(*_):
            path = cert_var.get().strip()
            ext  = os.path.splitext(path)[1].lower()
            is_pfx = ext in ('.pfx', '.p12')
            # Toggle key row
            state = tk.DISABLED if is_pfx else tk.NORMAL
            key_e.config(state=state, bg="#1e2a3a" if is_pfx else "#2a3f55")
            key_btn.config(state=state)
            key_lbl_var.set("Private key file  (not needed for PFX)"
                            if is_pfx else "Private key file  (.pem  .key)")
            # Toggle PFX password row
            if is_pfx:
                pfx_lbl.pack(anchor="w", pady=(8, 2))
                pfx_e.pack(fill="x", ipady=6)
            else:
                pfx_lbl.pack_forget()
                pfx_e.pack_forget()

        cert_var.trace_add("write", _on_cert_change)
        tk.Button(cert_row, text="Browse", bg="#2a5e8a", fg="white",
                  relief="flat", cursor="hand2", font=("Segoe UI", 9),
                  command=lambda: cert_var.set(
                      fd.askopenfilename(title="Select Certificate File",
                                         filetypes=CERT_TYPES) or cert_var.get())
                  ).pack(side="left", padx=(4, 0), ipadx=8, ipady=6)

        err_lbl = tk.Label(dlg, text="", bg="#1e2a3a", fg="#f44336", font=("Segoe UI", 9))
        err_lbl.pack(pady=(6, 0))

        def _restart_server(action_done):
            """Restart Flask to apply the new SSL config.
            - If Windows Service is running → stop + start the service.
            - If plain GUI mode → restart this process (os.execv).
            """
            import subprocess as _sp

            svc_running = False
            if HAS_WIN32SVC:
                try:
                    r = _sp.run(["sc", "query", "AssetManagerServer"],
                                capture_output=True, timeout=5)
                    svc_running = "RUNNING" in r.stdout.decode(errors="replace")
                except Exception:
                    pass

            if svc_running:
                mb.showinfo("Done", f"{action_done}\n\n"
                            "Restarting Windows Service to apply changes…\n"
                            "(Stats will refresh in a few seconds.)")
                def _do():
                    try:
                        _sp.run(["net", "stop", "AssetManagerServer"],
                                capture_output=True, timeout=30)
                        _sp.Popen(["net", "start", "AssetManagerServer"],
                                  stdout=_sp.DEVNULL, stderr=_sp.DEVNULL)
                    except Exception:
                        pass
                threading.Thread(target=_do, daemon=True).start()
            else:
                mb.showinfo("Done", f"{action_done}\n\n"
                            "Restarting server now to apply changes…")
                try:
                    os.execv(sys.executable, [sys.executable] + sys.argv)
                except Exception:
                    mb.showinfo("Restart Required",
                                "Please close and reopen AssetServer.exe to apply changes.")

        def do_install():
            c   = cert_var.get().strip()
            k   = key_var.get().strip()
            pwd = pfx_pwd_var.get()
            if not c:
                err_lbl.config(text="✗  Certificate file is required.")
                return
            if not os.path.exists(c):
                err_lbl.config(text="✗  Certificate file not found.")
                return
            try:
                cert_pem, key_pem = _cert_to_pem(c, k or None, pwd or None)
                with open(SSL_CERT_PATH, 'wb') as f:
                    f.write(cert_pem)
                with open(SSL_KEY_PATH, 'wb') as f:
                    f.write(key_pem)
                cfg = _load_server_cfg()
                cfg['ssl_enabled'] = True
                with open(SERVER_CFG_PATH, 'w') as f:
                    json.dump(cfg, f, indent=2)
                dlg.destroy()
                _restart_server(
                    f"✓  Certificate installed.\n\nClients connect to:\nhttps://{socket.gethostname()}:{PORT}"
                )
            except Exception as ex:
                err_lbl.config(text=f"✗  {ex}")

        def do_remove():
            if not mb.askyesno("Remove SSL",
                               "Remove the SSL certificate and revert to HTTP?",
                               icon="warning"):
                return
            try:
                for p in [SSL_CERT_PATH, SSL_KEY_PATH]:
                    if os.path.exists(p):
                        os.remove(p)
                cfg = _load_server_cfg()
                cfg['ssl_enabled'] = False
                with open(SERVER_CFG_PATH, 'w') as f:
                    json.dump(cfg, f, indent=2)
                dlg.destroy()
                _restart_server("✓  Certificate removed.")
            except Exception as ex:
                err_lbl.config(text=f"✗  {ex}")

        btn_frm = tk.Frame(dlg, bg="#1e2a3a")
        btn_frm.pack(fill="x", padx=18, pady=(6, 12))
        tk.Button(btn_frm, text="🔒  Install Certificate",
                  bg="#1a7a4a", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", pady=9,
                  command=do_install,
                  activebackground="#145f38").pack(fill="x", pady=(0, 4))
        if _cert_ok:
            tk.Button(btn_frm, text="Remove SSL",
                      bg="#4a4a5a", fg="white", font=("Segoe UI", 9),
                      relief="flat", cursor="hand2", pady=7,
                      command=do_remove,
                      activebackground="#333344").pack(fill="x")

    tk.Button(root, text="🔒  SSL Certificate",
              bg="#1a7a4a", fg="white",
              font=("Segoe UI", 9, "bold"),
              relief="flat", cursor="hand2", pady=8,
              command=upload_ssl_cert,
              activebackground="#145f38").pack(fill="x", padx=18, pady=(0, 10))

    # Stats
    stats_card = tk.Frame(root, bg="#243447", padx=14, pady=12)
    stats_card.pack(fill="x", padx=18, pady=(0, 18))
    tk.Label(stats_card, text="Live Statistics",
             bg="#243447", fg="#0078d4",
             font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
    stat_row   = tk.Frame(stats_card, bg="#243447")
    stat_row.pack(fill="x")
    stat_labels = {}
    for key, label, color in [
        ("total",     "Total",     "#ffffff"),
        ("pending",   "Pending",   "#f0a500"),
        ("confirmed", "Confirmed", "#4caf50"),
        ("out_now",   "Out Now",   "#f44336"),
    ]:
        col = tk.Frame(stat_row, bg="#2d3f52", padx=12, pady=8)
        col.pack(side="left", expand=True, fill="x", padx=3)
        v = tk.Label(col, text="—", bg="#2d3f52", fg=color,
                     font=("Segoe UI", 18, "bold"))
        v.pack()
        tk.Label(col, text=label, bg="#2d3f52", fg="#e0eaf5",
                 font=("Segoe UI", 10, "bold")).pack(pady=(4, 4))
        stat_labels[key] = v

    def refresh_stats():
        try:
            import requests as req, ssl as _ssl
            from requests.adapters import HTTPAdapter as _HA
            _ssl_on = (_SERVER_CFG.get('ssl_enabled') and
                       os.path.exists(SSL_CERT_PATH) and os.path.exists(SSL_KEY_PATH))
            _proto  = "https" if _ssl_on else "http"
            if _ssl_on:
                class _A(_HA):
                    def init_poolmanager(self, *a, **kw):
                        ctx = _ssl.SSLContext(_ssl.PROTOCOL_TLS_CLIENT)
                        ctx.check_hostname = False
                        ctx.verify_mode = _ssl.CERT_REQUIRED
                        if os.path.exists(SSL_CERT_PATH):
                            ctx.load_verify_locations(SSL_CERT_PATH)
                        kw['ssl_context'] = ctx
                        return super().init_poolmanager(*a, **kw)
                _s = req.Session(); _s.mount("https://", _A())
                r = _s.get(f"https://127.0.0.1:{PORT}/stats", timeout=2)
            else:
                r = req.get(f"http://127.0.0.1:{PORT}/stats", timeout=2)
            if r.ok:
                for k, lbl in stat_labels.items():
                    lbl.config(text=str(r.json().get(k, "—")))
        except: pass
        root.after(5000, refresh_stats)

    root.after(1500, refresh_stats)   # first poll after 1.5 s — Flask is guaranteed ready

    # Update tray tooltip periodically to reflect SSL state
    def _update_tray_title():
        if _tray_icon[0]:
            _ssl_now = _SERVER_CFG.get('ssl_enabled') and os.path.exists(SSL_CERT_PATH)
            _proto_now = "HTTPS" if _ssl_now else "HTTP"
            _tray_icon[0].title = f"Asset Manager — Server  ●  RUNNING  ({_proto_now})"
        root.after(10000, _update_tray_title)
    _update_tray_title()

    root.mainloop()


# ─── Flask runner (shared by interactive mode and Windows Service) ────────────

_AUTO_JOBS_STARTED = [False]

def _auto_accept_stale_pending():
    """Background job: auto-confirm transactions left pending for more than 24h."""
    while True:
        try:
            cutoff = (datetime.now() - timedelta(hours=24)).isoformat()
            conn = get_db()
            conn.execute(
                "UPDATE transactions SET confirmed=1, confirmed_at=? "
                "WHERE confirmed=0 AND rejected=0 AND timestamp < ?",
                (datetime.now().isoformat(), cutoff)
            )
            conn.commit()
            conn.close()
        except Exception:
            pass
        time.sleep(3600)


def _start_background_jobs():
    if _AUTO_JOBS_STARTED[0]:
        return
    _AUTO_JOBS_STARTED[0] = True
    threading.Thread(target=_auto_accept_stale_pending, daemon=True).start()


def _run_flask():
    """Start the Flask server (blocking). Chooses HTTP or HTTPS automatically."""
    _start_background_jobs()
    ssl_ctx = None
    if (_SERVER_CFG.get('ssl_enabled') and
            os.path.exists(SSL_CERT_PATH) and os.path.exists(SSL_KEY_PATH)):
        ssl_ctx = (SSL_CERT_PATH, SSL_KEY_PATH)
    app.run(
        host=HOST,
        port=PORT,
        debug=False,
        use_reloader=False,
        threaded=True,          # serve clients concurrently — one stuck/slow
                                # connection can't freeze the whole server
        ssl_context=ssl_ctx
    )


# ─── Windows Service definition ───────────────────────────────────────────────

if HAS_WIN32SVC:
    class AssetManagerService(win32serviceutil.ServiceFramework):
        """
        Windows Service wrapper for the Asset Manager Flask backend.
        Install : AssetServer.exe install
        Remove  : AssetServer.exe remove
        Start   : net start AssetManagerServer   (or via Services MMC)
        Stop    : net stop  AssetManagerServer
        Debug   : AssetServer.exe debug          (runs in console, Ctrl-C to stop)
        """
        _svc_name_         = "AssetManagerServer"
        _svc_display_name_ = "Asset Manager Server"
        _svc_description_  = (
            "Runs the Asset Manager REST API (Flask/SQLite) as a background service, "
            "independent of any logged-on user session."
        )

        def __init__(self, args):
            win32serviceutil.ServiceFramework.__init__(self, args)
            self._stop_evt = win32event.CreateEvent(None, 0, 0, None)

        def SvcStop(self):
            self.ReportServiceStatus(win32service.SERVICE_STOP_PENDING)
            win32event.SetEvent(self._stop_evt)

        def SvcDoRun(self):
            # Tell the SCM the service is now running — must be called before
            # any slow initialisation, otherwise net start times out (30 s default)
            self.ReportServiceStatus(win32service.SERVICE_RUNNING)
            servicemanager.LogMsg(
                servicemanager.EVENTLOG_INFORMATION_TYPE,
                servicemanager.PYS_SERVICE_STARTED,
                (self._svc_name_, ""),
            )
            init_db()
            _ensure_sample_lookup()
            threading.Thread(target=_run_flask, daemon=True).start()
            # Block until SvcStop sets the event
            win32event.WaitForSingleObject(self._stop_evt, win32event.INFINITE)
            servicemanager.LogMsg(
                servicemanager.EVENTLOG_INFORMATION_TYPE,
                servicemanager.PYS_SERVICE_STOPPED,
                (self._svc_name_, ""),
            )


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    # ── Service management commands (install / remove / start / stop / debug) ──
    _svc_cmds = {"install", "remove", "start", "stop", "restart", "debug",
                 "update", "--startup", "querycontrol"}
    _first_arg = sys.argv[1].lower() if len(sys.argv) > 1 else ""

    if _first_arg in _svc_cmds:
        if HAS_WIN32SVC:
            win32serviceutil.HandleCommandLine(AssetManagerService)
        else:
            print("pywin32 is not installed — Windows Service mode unavailable.")
            sys.exit(1)
    else:
        # ── Interactive GUI mode ──────────────────────────────────────────────
        # Hide the console window that appears when built without --windowed
        if sys.platform == "win32" and getattr(sys, 'frozen', False):
            try:
                ctypes.windll.user32.ShowWindow(
                    ctypes.windll.kernel32.GetConsoleWindow(), 0  # SW_HIDE
                )
            except Exception:
                pass

        if not _acquire_single_instance_mutex():
            # Another GUI window is already open — bring it to the foreground
            ctypes.windll.user32.MessageBoxW(
                0,
                "Asset Manager Server is already running.\n\nCheck the system tray.",
                "Already Running",
                0x40,   # MB_ICONINFORMATION
            )
            sys.exit(0)

        init_db()
        _ensure_sample_lookup()

        if _port_in_use(PORT):
            # Service (or another process) is already listening — skip Flask start
            # so the GUI acts as a monitor-only front-end
            pass
        else:
            threading.Thread(target=_run_flask, daemon=True).start()
            time.sleep(1.5)  # give Flask enough time to bind before the GUI polls it

        run_gui()
